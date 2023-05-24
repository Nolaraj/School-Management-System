from tkinter import scrolledtext
from tkinter import *
from tkinter import ttk
import datetime
from openpyxl import *
import os


class Clock:
    def __init__(self, bottom_frame):
        self.time1 = ''
        self.time2 = datetime.datetime.now().strftime('%H:%M:%S')
        self.mFrame = Frame(bottom_frame)
        self.mFrame.pack()

        self.watch = Label(self.mFrame, text=self.time2, font=('verdana', 12, 'italic'))
        self.watch.pack()

        self.changeLabel()  # first call it manually

    def changeLabel(self):
        self.time2 = datetime.datetime.now().strftime('%H:%M:%S')
        self.watch.configure(text=self.time2)
        self.mFrame.after(200, self.changeLabel)  # it'll call itself continuously


root = Tk()

sizex = 1300
sizey = 630
posx = 0
posy = 0
school_name = "ABC Boarding School"
counter = IntVar()
counter = 1
# navigation_image = PhotoImage(file="bg_navigation.png")
# welcome_image = PhotoImage(file="school__photo.png")

# ------------

main_path = os.getcwd()
main_folder = "School Database"
path1 = main_path + "\\" + main_folder
folder1 = "Students Database"
folder2 = "Teachers Database"
folder3 = "Staffs Database"
folder4 = "Utilities"
path11 = path1 + "\\" + folder1
path12 = path1 + "\\" + folder2
path13 = path1 + "\\" + folder3
path_utilities = path1 + "\\" + folder4
folder11 = "Composite Classwise Records"
folder12 = "Individual Records"
folder13 = "Registration Records"
folder14 = "MixedUp Records"
path111 = path11 + "\\" + folder11
path112 = path11 + "\\" + folder12
path113 = path11 + "\\" + folder13
path114 = path11 + "\\" + folder14
folder12pnur = "Pre-Nursery"
folder12nur = "Nursery"
folder12lkg = "LKG"
folder12ukg = "UKG"
folder121 = "Class 1"
folder122 = "Class 2"
folder123 = "Class 3"
folder124 = "Class 4"
folder125 = "Class 5"
folder126 = "Class 6"
folder127 = "Class 7"
folder128 = "Class 8"
folder129 = "Class 9"
folder1210 = "Class 10"

path112pn = path112 + "\\" + folder12pnur
path112n = path112 + "\\" + folder12nur
path112l = path112 + "\\" + folder12lkg
path112u = path112 + "\\" + folder12ukg
path1121 = path112 + "\\" + folder121
path1122 = path112 + "\\" + folder122
path1123 = path112 + "\\" + folder123
path1124 = path112 + "\\" + folder124
path1125 = path112 + "\\" + folder125
path1126 = path112 + "\\" + folder126
path1127 = path112 + "\\" + folder127
path1128 = path112 + "\\" + folder128
path1129 = path112 + "\\" + folder129
path11210 = path112 + "\\" + folder1210

folder21 = "Composite  Records"
folder22 = "Individual Records"
path121 = path12 + "\\" + folder21
path122 = path12 + "\\" + folder22
folder31 = "Composite  Records"
folder32 = "Individual Records"
path131 = path13 + "\\" + folder31
path132 = path13 + "\\" + folder32

folder_reg = "Courses Registration"
path_reg = path1 + "\\" + folder_reg
# ---------------------------------------------

year_today = datetime.datetime.now().strftime('%Y')
month_today = datetime.datetime.now().strftime('%m')
day_today = datetime.datetime.now().strftime('%d')
hour_today = datetime.datetime.now().strftime('%H')
minutes_today = datetime.datetime.now().strftime('%M')
seconds_today = datetime.datetime.now().strftime('%S')


def date_time1():
    date_time = datetime.datetime.now().strftime('%Y') + "/" + datetime.datetime.now().strftime \
        ('%m') + "/" + datetime.datetime.now().strftime('%d') + "//" + datetime.datetime.now().strftime \
                    ('%H') + ":" + datetime.datetime.now().strftime('%M') + ":" + datetime.datetime.now().strftime('%S')
    return date_time
def date():
    date = datetime.datetime.now().strftime('%Y') + "/" + datetime.datetime.now().strftime \
        ('%m') + "/" + datetime.datetime.now().strftime('%d')
    return date
def time():
    time = datetime.datetime.now().strftime \
                    ('%H') + ":" + datetime.datetime.now().strftime('%M') + ":" + datetime.datetime.now().strftime('%S')
    return time


# ------------ Main Window Assignment
root.geometry("%dx%d+%d+%d" % (sizex, sizey, posx, posy))
root.state('zoomed')
root.title("School Management System")


def framework(root):
    def log_handler(text1, text2, text3, text4, text5, text6):
        # -------------------------------------------------------
        os.chdir(path_utilities)
        log_filename = "Log File.xlsx"
        log_sheetname = "Log"
        try:
            log_book = load_workbook(log_filename)
            log_sheet = log_book(log_sheetname)
        except:
            log_book = Workbook()
            log_sheet = log_book.active
            log_sheet.title = log_sheetname
            log_sheet.cell(row=3, column=1).value = "S.N."
            log_sheet.cell(row=3, column=2).value = "Database Class"
            log_sheet.cell(row=3, column=3).value = "Event Title"
            log_sheet.cell(row=3, column=4).value = "Activity Type"
            log_sheet.cell(row=3, column=5).value = "Date"
            log_sheet.cell(row=3, column=6).value = "Time"
            log_sheet.cell(row=3, column=7).value = "Registered By"
        row_number = log_sheet.max_row
        log_sheet.cell(row=row_number+1, column=1).value = row_number - 3
        log_sheet.cell(row=row_number+1, column=2).value = text1
        log_sheet.cell(row=row_number+1, column=3).value = text2
        log_sheet.cell(row=row_number+1, column=4).value = text3
        log_sheet.cell(row=row_number+1, column=5).value = text4
        log_sheet.cell(row=row_number+1, column=6).value = text5
        log_sheet.cell(row=row_number+1, column=7).value = text6




        # -------------------------------------------------------
    def myfunction_button1(event):
        canvas.configure(scrollregion=canvas.bbox("all"), width=sizex - 50, height=sizey)

    def myfunction_button11(event):
        maincanvas.configure(scrollregion=maincanvas.bbox("all"), width=695, height=460)

    def myfunction_button12(event):
        maincanvas.configure(scrollregion=maincanvas.bbox("all"), width=695, height=470)

    def date_converter_returning_age(provided_year, provided_month1, provided_day, now_year, now_month, now_day):
        month1 = ("Baisakh", "Jestha", "Ashadh", "Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Paush",
                  "Magh", "Falgun", "Chaitra")
        month2 = ("1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12")
        month_n = {}
        month_n = dict(zip(month1, month2))

        if (((provided_month1 == 1 or provided_month1 == 2) or (provided_month1 == 3 or provided_month1 == 4)) or
            ((provided_month1 == 5 or provided_month1 == 6) or (provided_month1 == 7 or provided_month1 == 8))) or (
                (provided_month1 == 9 or provided_month1 == 10) or (provided_month1 == 11 or provided_month1 == 11)):
            provided_month = provided_month1
        else:
            provided_month = month_n.get(provided_month1)

        print(provided_month)

        year = int(provided_year)
        month = int(provided_month)
        day = int(provided_day)
        a = int(now_year)
        b = int(now_month)
        c = int(now_day)
        if month < 10:
            if month == 9 and day > 15:
                day -= 15
                month = 1
                year -= 56
            elif month == 9 and (day < 15 or day == 15):
                day += 15
                month = 12
                year -= 57
            elif (day < 15 or day == 15):
                day += 15
                month += 3
                year -= 57
            else:
                day -= 15
                month += 4
                year -= 57
        else:
            year -= 56
            if (day < 15 or day == 15):
                day += 15
                month -= 9

            else:
                day -= 15
                month -= 8

        # Age Calculation Stars
        if b > month:
            if c > day:
                year1 = a - year
                month1 = b - month
                day1 = c - day
            else:
                year1 = a - year
                month1 = (b - month) - 1
                day1 = 30 - (day - c)

        elif b < month:
            if c > day:
                year1 = (a - year) - 1
                month1 = 12 - (month - b)
                day1 = c - day
            else:
                day1 = 30 - (day - c)
                month1 = 11 - (month - b)
                year1 = (a - year) - 1
        elif c == day:
            if b > month:
                day1 = 0
                month1 = b - month
                year1 = a - year
            else:
                day1 = 0
                month1 = 12 - (month - b)
                year1 = (a - year) - 1
        else:
            if c > day:
                day1 = c - day
                month1 = b - month
                year1 = a - year
            else:
                day1 = 30 - (day - c)
                month1 = 11
                year1 = a - year

        age = str(year1) + " " + "Years" + "  " + str(month1) + " " + "Month" + "  " + str(day1) + " " + "Days"

        return age

    def new_registration_event(event):
        def student_to_db(event):
            date_time = date_time1()
            s0 = "Studying"
            s1 = student_name.get()
            s2 = student_class.get()
            s3 = student_gender.get()
            s4 = student_dob1.get()
            s4a = student_dob2.get()
            s4b = student_dob3.get()
            s5 = student_paddress.get()
            s5a = student_paddress1.get()
            s5b = student_paddress2.get()
            s6 = student_taddress.get()
            s6a = student_taddress1.get()
            s6b = student_taddress2.get()
            s7 = student_father.get()
            s8 = student_mother.get()
            s9 = student_gfather.get()
            s10 = student_eyear.get()
            s11 = student_pschool.get()
            s12 = student_accomodation.get()
            sc = str(student_eyear.get() % 100) + student_gfather.get()[0] + student_father.get()[0] + \
                 student_name.get()[0]
            composite_datas = [sc, date_time, s0, s1, s2, s3, s4, s4a, s4b, s5, s5a, s5b, s6, s6a, s6b, s7, s8, s9, s10,
                               s11, s12]

            titles = ["S.N.", "Student Code", "Date/Time of Entry", "Current Status", "Name of the Student", "Class",
                      "Gender", "Date of Birth", " ", " ", "Permanent Location", " ", " ", "Temporary Location", " ",
                      " ", "Father's Name", "Mother's Name", "Grand Father's Name", "Enrollment Year",
                      "Previous School", "Need of Hostel's Accomodation?"]
            dob = ["Year", "Month", "Day"]
            location = ["District", "Local Government", "Area/Tol"]

            def individual_writer():
                def write_class(cell_value, cell_no, extra, numb):
                    if cell_value != "":
                        c_v = "D" + str(cell_no)
                        worksheet1[c_v].value = extra
                        if numb == 2:
                            c_v = "E" + str(cell_no)
                            worksheet1[c_v].value = cell_value

                filename = str(student_eyear.get() % 100) + student_gfather.get()[0] + student_father.get()[0] + \
                           student_name.get()[0] + " " + student_name.get()
                workbook1 = Workbook()
                worksheet1 = workbook1.active
                worksheet1.title = student_name.get()

                i = 4
                j = 4
                data_title = ["Name of the Student", "Class", "Gender", "Date of Birth", "Permanent Address",
                              "Temporary Address",
                              "Father's Name", "Mother's Name", "Grand Father's Name", "Year of Enrollment",
                              "Previous School's Name",
                              "Need of Hostel's Accomodation?"]
                entry_datas = [s1, s2, s3, s4, s5, s6, s7, s8, s9, s10, s11, s12]
                for item in data_title:
                    row = i
                    s_cell = "B" + str(row)
                    worksheet1[s_cell].value = item
                    i += 1
                for item in entry_datas:
                    row = j
                    s_cell = "C" + str(row)
                    worksheet1[s_cell].value = item
                    j += 1
                write_class(s4b, 7, s4a, 2)
                write_class(s5b, 8, s5a, 2)
                write_class(s6b, 9, s6a, 2)
                worksheet1["B1"] = "Created Date"
                worksheet1["B2"] = year_today + "/" + month_today + "/" + day_today
                worksheet1["D1"] = "Created Time"
                worksheet1["D2"] = hour_today + "/" + minutes_today + "/" + seconds_today
                worksheet1["B3"] = "Current Status"
                worksheet1["C3"] = "Studying"

                if student_class.get() == "Pre-Nursery":
                    fixed_path = path112pn
                elif student_class.get() == "Nursery":
                    fixed_path = path112n
                elif student_class.get() == "LKG":
                    fixed_path = path112l
                elif student_class.get() == "UKG":
                    fixed_path = path112u
                elif student_class.get() == "1":
                    fixed_path = path1121
                elif student_class.get() == "2":
                    fixed_path = path1122
                elif student_class.get() == "3":
                    fixed_path = path1123
                elif student_class.get() == "4":
                    fixed_path = path1124
                elif student_class.get() == "5":
                    fixed_path = path1125
                elif student_class.get() == "6":
                    fixed_path = path1126
                elif student_class.get() == "7":
                    fixed_path = path1127
                elif student_class.get() == "8":
                    fixed_path = path1128
                elif student_class.get() == "9":
                    fixed_path = path1129
                elif student_class.get() == "10":
                    fixed_path = path11210

                filename = f'{filename}.xlsx'
                workbook1.save(os.path.join(fixed_path, filename))

            def composite_writer_cw():
                os.chdir(path111)
                filename = "Class" + " " + s2

                try:
                    workbook2 = load_workbook(filename=f'{filename}.xlsx')
                except:
                    workbook2 = Workbook()
                if filename in workbook2.sheetnames:
                    worksheet2 = workbook2[filename]
                else:
                    worksheet2 = workbook2.active
                    worksheet2.title = filename
                if worksheet2.max_row == 1:
                    for column in range(1, 23):
                        for row in range(4, 5):
                            worksheet2.cell(row=row, column=column).value = titles[column - 1]
                    for column in range(8, 11):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = dob[column - 8]
                    for column in range(11, 14):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = location[column - 11]
                    for column in range(14, 17):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = location[column - 14]
                    composite_datas.insert(0, 1)
                    for column in range(1, 23):
                        for row in range(6, 7):
                            worksheet2.cell(row=row, column=column).value = composite_datas[column - 1]
                    composite_datas.pop(0)
                else:
                    composite_datas.insert(0, worksheet2.max_row - 4)
                    row1 = worksheet2.max_row + 1
                    row2 = worksheet2.max_row + 2
                    for column in range(1, 23):
                        for row in range(row1, row2):
                            worksheet2.cell(row=row, column=column).value = composite_datas[column - 1]
                    composite_datas.pop(0)
                worksheet2['B1'] = "Last Modified On:"
                worksheet2['C1'] = date_time

                workbook2.save(f'{filename}.xlsx')

            def composite_writer_mp():
                os.chdir(path114)
                filename = "All Students"

                try:
                    workbook2 = load_workbook(filename=f'{filename}.xlsx')
                except:
                    workbook2 = Workbook()
                if filename in workbook2.sheetnames:
                    worksheet2 = workbook2[filename]
                else:
                    worksheet2 = workbook2.active
                    worksheet2.title = filename
                if worksheet2.max_row == 1:
                    for column in range(1, 23):
                        for row in range(4, 5):
                            worksheet2.cell(row=row, column=column).value = titles[column - 1]
                    for column in range(8, 11):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = dob[column - 8]
                    for column in range(11, 14):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = location[column - 11]
                    for column in range(14, 17):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = location[column - 14]
                    composite_datas.insert(0, 1)
                    for column in range(1, 23):
                        for row in range(6, 7):
                            worksheet2.cell(row=row, column=column).value = composite_datas[column - 1]
                    composite_datas.pop(0)
                else:
                    composite_datas.insert(0, worksheet2.max_row - 4)
                    row1 = worksheet2.max_row + 1
                    row2 = worksheet2.max_row + 2
                    for column in range(1, 23):
                        for row in range(row1, row2):
                            worksheet2.cell(row=row, column=column).value = composite_datas[column - 1]
                    composite_datas.pop(0)
                worksheet2['B1'] = "Last Modified On:"
                worksheet2['C1'] = date_time

                workbook2.save(f'{filename}.xlsx')

            def composite_writer_rg():
                os.chdir(path113)
                filename = f'Registration-{s10}'

                try:
                    workbook2 = load_workbook(filename=f'{filename}.xlsx')
                except:
                    workbook2 = Workbook()
                if filename in workbook2.sheetnames:
                    worksheet2 = workbook2[filename]
                else:
                    worksheet2 = workbook2.active
                    worksheet2.title = filename
                if worksheet2.max_row == 1:
                    for column in range(1, 23):
                        for row in range(4, 5):
                            worksheet2.cell(row=row, column=column).value = titles[column - 1]
                    for column in range(8, 11):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = dob[column - 8]
                    for column in range(11, 14):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = location[column - 11]
                    for column in range(14, 17):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = location[column - 14]
                    composite_datas.insert(0, 1)
                    for column in range(1, 23):
                        for row in range(6, 7):
                            worksheet2.cell(row=row, column=column).value = composite_datas[column - 1]
                    composite_datas.pop(0)
                else:
                    composite_datas.insert(0, worksheet2.max_row - 4)
                    row1 = worksheet2.max_row + 1
                    row2 = worksheet2.max_row + 2
                    for column in range(1, 22):
                        for row in range(row1, row2):
                            worksheet2.cell(row=row, column=column).value = composite_datas[column - 1]
                    composite_datas.pop(0)
                worksheet2['B1'] = "Last Modified On:"
                worksheet2['C1'] = date_time
                workbook2.save(f'{filename}.xlsx')
                text = f"{str(student_eyear.get() % 100)}{student_gfather.get()[0]}{student_father.get()[0]}{student_name.get()[0]} {student_name.get()}.xlsx  File has been created in {date_time} \n"
                act_text.configure(state='normal')
                act_text.insert('1.0', f'\n {text}')
                act_text.configure(state='disabled')
                text1 = f"Creation of {str(student_eyear.get() % 100)}{student_gfather.get()[0]}{student_father.get()[0]}{student_name.get()[0]} {student_name.get()}.xlsx"
                print(date, time)
                log_handler("Students Database", text1, "New Registration", date, time, "Anonymous")

            individual_writer()
            composite_writer_cw()
            composite_writer_mp()
            composite_writer_rg()

        def clear_all(event):
            student_name1.delete(0, "end")
            student_class1.delete(0, "end")
            student_gender1.delete(0, "end")
            student_dob11.delete(0, "end")
            student_dob12.delete(0, "end")
            student_dob13.delete(0, "end")
            student_plocation11.delete(0, "end")
            student_plocation12.delete(0, "end")
            student_plocation13.delete(0, "end")
            student_tlocation21.delete(0, "end")
            student_tlocation22.delete(0, "end")
            student_tlocation23.delete(0, "end")
            student_fname.delete(0, "end")
            student_mother1.delete(0, "end")
            student_gfname.delete(0, "end")
            student_enroll.delete(0, "end")
            student_pschool1.delete(0, "end")
            student_haccomodation.delete(0, "end")

        # todo After Some addtion in the components of the Student
        def preview(event):
            teacher_preview1.configure(state='normal')

            teacher_preview1.delete('1.0', "end-1c")
            teacher_preview1.insert(INSERT, f' \t Summary of the Provided Information \n')
            teacher_preview1.insert(INSERT, f' Personal Information \n')
            teacher_preview1.insert(INSERT, f' Name \t  {teacher_name.get()} \n')
            teacher_preview1.insert(INSERT, f' Age \t  {teacher_age.get()} \n')
            teacher_preview1.insert(INSERT, f' Gender \t  {teacher_gender.get()} \n')
            teacher_preview1.insert(INSERT,
                                    f' Date of Birth \t  {teacher_dob1.get()} \t {teacher_dob2.get()} \t {teacher_dob3.get()} \n')
            teacher_preview1.insert(INSERT, f' Marriage Status \t  {teacher_marriage.get()} \n')
            teacher_preview1.insert(INSERT,
                                    f' Permanent Address \t  {teacher_paddress.get()}  \t {teacher_paddress1.get()} \t {teacher_paddress2.get()} \n')
            teacher_preview1.insert(INSERT,
                                    f' Temporary Address \t  {teacher_taddress.get()} \t {teacher_taddress1.get()} \t {teacher_taddress2.get()}\n')
            teacher_preview1.insert(INSERT, f' Contact No. \t  {teacher_contact1.get()} \n')
            teacher_preview1.insert(INSERT, f' Secondary Contact No. \t  {teacher_contact2.get()} \n')
            teacher_preview1.insert(INSERT, f' E-Mail Address \t  {teacher_email.get()} \n')
            teacher_preview1.insert(INSERT, f' Facebook ID \t  {teacher_facebook.get()} \n')
            teacher_preview1.insert(INSERT, f'\n Education \t  \n')
            teacher_preview1.insert(INSERT, f' Level \t  {teacher_level1.get()} \n')
            teacher_preview1.insert(INSERT, f' Board \t  {teacher_board1.get()} \n')
            teacher_preview1.insert(INSERT, f' Institution \t  {teacher_institution1.get()} \n')
            teacher_preview1.insert(INSERT, f' Grade/Percentage \t  {teacher_grade1.get()} \n')
            teacher_preview1.insert(INSERT, f' Main Subject \t  {teacher_subject.get()} \n')
            teacher_preview1.insert(INSERT, f'\n Experiences and Eligibility \t \n')
            teacher_preview1.insert(INSERT, f' Field of Work \t  {teacher_fow1.get()}  \n')
            teacher_preview1.insert(INSERT, f' Post \t  {teacher_post1.get()} \n')
            teacher_preview1.insert(INSERT, f' Duration \t  {teacher_duration1.get()} \t Year/s\n')
            teacher_preview1.insert(INSERT, f' Institution \t  {teacher_wintitution1.get()} \n')
            teacher_preview1.insert(INSERT, f' Location \t  {teacher_wlocation1.get()} \n')
            teacher_preview1.insert(INSERT, f'\t  \n')
            teacher_preview1.insert(INSERT, f' Description \t  {teacher_description} \n')
            teacher_preview1.insert(INSERT, f'\n Teaching Subject Entry \t  \n')
            teacher_preview1.insert(INSERT, f' Main Subject \t  {teacher_mainsubject.get()}')
            teacher_preview1.insert(INSERT, f' For Class/es \t {teacher_class1.get()} ')
            teacher_preview1.insert(INSERT, f' To \t {teacher_class2.get()} \n')
            teacher_preview1.insert(INSERT, f' Secondary Subject1 \t  {teacher_ssubject1.get()} ')
            teacher_preview1.insert(INSERT, f' For Class/es\t {teacher_class21.get()} ')
            teacher_preview1.insert(INSERT, f' To \t {teacher_class22.get()} \n')
            teacher_preview1.insert(INSERT, f' Secondary Subject2 \t  {teacher_ssubject2.get()} ')
            teacher_preview1.insert(INSERT, f' For Class/es \t {teacher_class31.get()} ')
            teacher_preview1.insert(INSERT, f' To \t {teacher_class32.get()} \n')

            teacher_preview1.configure(state='disabled')

        # ------------
        student_name = StringVar()
        student_class = StringVar()
        student_gender = StringVar()
        student_dob1 = StringVar()
        student_dob2 = StringVar()
        student_dob3 = StringVar()
        student_paddress = StringVar()
        student_paddress1 = StringVar()
        student_paddress2 = StringVar()
        student_taddress = StringVar()
        student_taddress1 = StringVar()
        student_taddress2 = StringVar()
        student_father = StringVar()
        student_mother = StringVar()
        student_gfather = StringVar()
        student_eyear = IntVar()
        student_pschool = StringVar()
        student_accomodation = StringVar()

        options_age = list(range(15, 80))
        options_gender = ("Male", "Female", "Other")
        options_district = (
            "Bhaktapur", "Chitwan", "Dhading", "Dolakha", "Kathmandu", "Kavrepalanchok", "Lalitpur", "Makwanpur",
            "Nuwakot",
            "Ramechhap", "Rasuwa", "Sindhuli", "Sindhupalchok", "Bhojpur", "Dhankuta", "Ilam", "Jhapa", "Khotang",
            "Morang",
            "Okhaldhunga", "Panchthar", "Sankhuwasabha", "Solukhumbu", "Sunsari", "Taplejung", "Terhathum",
            "Udayapur",
            "Bara",
            "Dhanusa", "Mahottari", "Parsa", "Rautahat", "Saptari", "Sarlahi", "Siraha", "Baglung", "Gorkha",
            "Kaski",
            "Lamjung", "Manang", "Mustang", "Myagdi", "Nawalpur", "Parbat", "Syangja", "Tanahun", "Arghakhanchi",
            "Banke",
            "Bardiya", "Dang", "Eastern Rukum", "Gulmi", "Kapilvastu", "Palpa", "Parasi", "Pyuthan", "Rolpa",
            "Rupandehi",
            "Achham", "Baitadi", "Bajhang", "Bajura", "Dadeldhura", "Darchula", "Doti", "Kailali", "Kanchanpur",
            "Dailekh",
            "Dolpa", "Humla", "Jajarkot", "Jumla", "Kalikot", "Mugu", "Salyan", "Surkhet", "Western Rukum")
        options_local_level = (
            "Pokhara Metropolitan City", "Annapurna", "Machhapuchhre", "Madi", "Rupa", "Badigad", "Kathekhola",
            "Nisikhola",
            "Bareng", "Tarakhola", "Tamankhola", "Shahid Lakhan", "Barpak Sulikot", "Aarughat", "Siranchowk",
            "Gandaki",
            "Bhimsen Thapa", "Ajirkot", "Dharche", "Tsum Nubri", "Marsyangdi", "Dordi", "Dudhpokhari",
            "Kwaholasothar",
            "Manang Disyang", "Nason", "Chame", "Narpa Bhumi", "Gharapjhong", "Thasang", "Baragung Muktichhetra",
            "Lomanthang",
            "Lo-Ghekar Damodarkunda", "Malika", "Mangala", "Raghuganga", "Dhaulagiri", "Annapurna", "Hupsekot",
            "Binayi Triveni", "Bulingtar", "Baudikali", "Kaligandaki", "Biruwa", "Harinas", "Aandhikhola",
            "Arjun Chaupari",
            "Phedikhola", "Rishing", "Myagde", "Aanbu Khaireni", "Bandipur", "Ghiring", "Devghat")
        options_classes = (
            "Pre-Nursery", "Nursery", "LKG", "UKG", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10")

        options_years = (
            "2070", "2071", "2072", "2073", "2074", "2075", "2076", "2077", "2078", "2079", "2080", "2081", "2082",
            "2083",
            "2084", "2085", "2086", "2087", "2088")
        options_options = ("Yes", "No", "Not Specified")
        options_year = (
            "2078", "2077", "2076", "2075", "2074", "2073", "2072", "2071", "2070", "2069", "2068", "2067", "2066",
            "2065",
            "2064", "2063", "2062", "2061", "2060", "2059", "2058", "2057", "2056", "2055", "2054", "2053", "2052",
            "2051",
            "2050", "2049", "2048", "2047", "2046", "2045", "2044", "2043", "2042", "2041", "2040", "2039", "2038",
            "2037",
            "2036", "2035", "2034", "2033", "2032", "2031", "2030", "2029", "2028", "2027", "2026", "2025", "2024",
            "2023",
            "2022", "2021", "2020")
        options_month = (
            "Baisakh", "Jestha", "Ashadh", "Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Paush", "Magh",
            "Falgun",
            "Chaitra")
        options_day = (
            "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
            "20", "21",
            "22", "23", "24", "25", "26", "27", "28", "29", "30", "31", "32")

        registration_navigation = Frame(detail_pane_frame, relief=FLAT, bg="lightcyan", width=730, height=470)
        registration_navigation.grid(row=0, column=0, sticky='nsew')
        student_title_label = Label(registration_navigation, text="Register - Details of Student",
                                    font=("verdana", 12, 'bold'),
                                    relief=FLAT, fg="black", bg="coral", width=62)
        student_title_label.grid(row=0, columnspan=3, pady=10)
        student_name_label = Label(registration_navigation, text="Name of the Student", font=("verdana", 10),
                                   width=35,
                                   relief=FLAT, fg="black", bg="bisque")
        student_name_label.grid(row=1, column=0, pady=5, padx=5)
        student_name1 = Entry(registration_navigation, font=("verdana", 10), width=35, relief=SUNKEN, fg="black",
                              textvariable=student_name)
        student_name1.grid(row=1, column=1, pady=5)
        student_class_label = Label(registration_navigation, text="Class of Registration", font=("verdana", 10),
                                    width=35, relief=FLAT, fg="black", bg="bisque")
        student_class_label.grid(row=2, column=0, pady=5)
        student_class1 = ttk.Combobox(registration_navigation, state='readonly', values=options_classes,
                                      font=("verdana", 10), width=33, textvariable=student_class)
        student_class1.grid(row=2, column=1, pady=5)
        student_gender_label = Label(registration_navigation, text="Gender", font=("verdana", 10), width=35,
                                     relief=FLAT,
                                     fg="black", bg="bisque")
        student_gender_label.grid(row=3, column=0, pady=5)
        student_gender1 = ttk.Combobox(registration_navigation, state='readonly', values=options_gender,
                                       font=("verdana", 10),
                                       width=33, textvariable=student_gender)
        student_gender1.grid(row=3, column=1, pady=5)
        student_dob_label = Label(registration_navigation, text="Date of Birth(YYYY-MM-DD)", font=("verdana", 10),
                                  width=35, relief=FLAT, fg="black", bg="bisque")
        student_dob_label.grid(row=4, column=0, pady=5)
        student_dob10 = Frame(registration_navigation, width=35)
        student_dob10.grid(row=4, column=1, pady=5)
        student_dob11 = ttk.Combobox(student_dob10, state="readonly", values=options_year,
                                     font=("verdana", 10), width=9, textvariable=student_dob1)
        student_dob11.pack(side=LEFT)
        student_dob12 = ttk.Combobox(student_dob10, state="readonly", values=options_month,
                                     font=("verdana", 10), width=9, textvariable=student_dob2)
        student_dob12.pack(side=LEFT)
        student_dob13 = ttk.Combobox(student_dob10, state="readonly", values=options_day,
                                     font=("verdana", 10), width=9, textvariable=student_dob3)
        student_dob13.pack(side=LEFT)

        student_plocation_label = Label(registration_navigation, text="Permanent Location", font=("verdana", 10),
                                        width=35,
                                        relief=FLAT, fg="black", bg="bisque")
        student_plocation_label.grid(row=5, column=0, pady=5)
        student_plocation = Frame(registration_navigation, width=35, relief=SUNKEN)
        student_plocation.grid(row=5, column=1, pady=5)
        student_plocation11 = ttk.Combobox(student_plocation, state="readonly", values=options_district,
                                           font=("verdana", 10), width=9, textvariable=student_paddress)
        student_plocation11.current(37)
        student_plocation11.pack(side=LEFT)
        student_plocation12 = ttk.Combobox(student_plocation, state="readonly", values=options_local_level,
                                           font=("verdana", 10), width=9, textvariable=student_paddress1)
        student_plocation12.current(0)
        student_plocation12.pack(side=LEFT)
        student_plocation13 = Entry(student_plocation, font=("verdana", 10), width=11, relief=SUNKEN, fg="black",
                                    textvariable=student_paddress2)
        student_plocation13.pack(side=LEFT)
        student_tlocation_label = Label(registration_navigation, text="Temporary Location", font=("verdana", 10),
                                        width=35, relief=FLAT, fg="black", bg="bisque")
        student_tlocation_label.grid(row=6, column=0, pady=5)
        student_tlocation = Frame(registration_navigation, width=35, relief=SUNKEN)
        student_tlocation.grid(row=6, column=1, pady=5)
        student_tlocation21 = ttk.Combobox(student_tlocation, state="readonly", values=options_district,
                                           font=("verdana", 10), width=9, textvariable=student_taddress)
        student_tlocation21.current(37)
        student_tlocation21.pack(side=LEFT)
        student_tlocation22 = ttk.Combobox(student_tlocation, state="readonly", values=options_local_level,
                                           font=("verdana", 10), width=9, textvariable=student_taddress1)
        student_tlocation22.current(0)
        student_tlocation22.pack(side=LEFT)
        student_tlocation23 = Entry(student_tlocation, font=("verdana", 10), width=11, relief=SUNKEN, fg="black",
                                    textvariable=student_taddress2)
        student_tlocation23.pack(side=LEFT)
        student_fname_label = Label(registration_navigation, text="Fathers' Name", font=("verdana", 10), width=35,
                                    relief=FLAT, fg="black", bg="bisque")
        student_fname_label.grid(row=7, column=0, pady=5)
        student_fname = Entry(registration_navigation, font=("verdana", 10), width=35, relief=SUNKEN, fg="black",
                              textvariable=student_father)
        student_fname.grid(row=7, column=1, pady=5)
        student_mother_label = Label(registration_navigation, text="Mothers' Name", font=("verdana", 10),
                                     width=35, relief=FLAT, fg="black", bg="bisque")
        student_mother_label.grid(row=8, column=0, pady=5)
        student_mother1 = Entry(registration_navigation, font=("verdana", 10), width=35, relief=SUNKEN,
                                fg="black", textvariable=student_mother)
        student_mother1.grid(row=8, column=1, pady=5)
        student_gfname_label = Label(registration_navigation, text="Grand Fathers' Name", font=("verdana", 10),
                                     width=35,
                                     relief=FLAT, fg="black", bg="bisque")
        student_gfname_label.grid(row=9, column=0, pady=5)
        student_gfname = Entry(registration_navigation, font=("verdana", 10), width=35, relief=SUNKEN, fg="black",
                               textvariable=student_gfather)
        student_gfname.grid(row=9, column=1, pady=5)
        student_enroll_label = Label(registration_navigation, text="Enrollment Year", font=("verdana", 10),
                                     width=35,
                                     relief=FLAT, fg="black", bg="bisque")
        student_enroll_label.grid(row=10, column=0, pady=5)
        student_enroll = ttk.Combobox(registration_navigation, state="readonly", values=options_years,
                                      font=("verdana", 10), width=33, textvariable=student_eyear)
        student_enroll.current(8)
        student_enroll.grid(row=10, column=1, pady=5)
        student_pschool_label = Label(registration_navigation, text="Previous School", font=("verdana", 10),
                                      width=35,
                                      relief=FLAT, fg="black", bg="bisque")
        student_pschool_label.grid(row=11, column=0, pady=5)
        student_pschool1 = Entry(registration_navigation, font=("verdana", 10), width=35, relief=SUNKEN, fg="black",
                                 textvariable=student_pschool)
        student_pschool1.grid(row=11, column=1, pady=5)
        student_haccomodation_label = Label(registration_navigation, text="Hostels' Accomodation?",
                                            font=("verdana", 10),
                                            width=35, relief=FLAT, fg="black", bg="bisque")
        student_haccomodation_label.grid(row=12, column=0, pady=5)
        student_haccomodation = ttk.Combobox(registration_navigation, state="readonly", values=options_options,
                                             font=("verdana", 10),
                                             width=33, textvariable=student_accomodation)
        student_haccomodation.grid(row=12, column=1, pady=5)

        student_button_submit = Button(registration_navigation, text="Submit", font=("verdana", 12), relief=RAISED,
                                       bg="springgreen2", width=20)
        student_button_submit.bind("<Button-1>", student_to_db)
        student_button_submit.grid(row=13, column=0, pady=5)
        student_button_clear = Button(registration_navigation, text=f"Clear All", font=("verdana", 12),
                                      relief=RAISED,
                                      bg="lightpink", width=20)
        student_button_clear.bind("<Button-1>", clear_all)
        student_button_clear.grid(row=13, column=1, pady=5)

    def existing_record_event(event):
        search_type = StringVar()
        search_text = StringVar()

        existing_navigation = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
        existing_navigation.grid(row=0, column=0, sticky="nsew")

        options_search = (
            "", "Search by Name", "Search by Class", "Search by Registration Year", "Search by Permanent Address",
            "Search by Temporary Address", "Search by Father's Name")

        def search_student(event):
            os.chdir(path114)
            search_book = load_workbook("All Students.xlsx")
            search_sheet = search_book["All Students"]
            counter = 1
            text = search_text.get()

            search_frame = Frame(detail_pane_frame)
            search_frame.grid(row=0, column=0, sticky="nsew")

            search_frame1 = Frame(search_frame)
            search_frame1.grid(row=0, column=0, sticky="ew")

            search_frame2 = Frame(search_frame)
            search_frame2.grid(row=1, column=0, sticky="nsew")

            frame_searched = Frame(search_frame2)
            frame_searched.pack(side=TOP, fill=X)

            def button_creator(student_classe, filename, counter):

                if student_classe == "Pre-Nursery":
                    fixed_path = path112pn
                elif student_classe == "Nursery":
                    fixed_path = path112n
                elif student_classe == "LKG":
                    fixed_path = path112l
                elif student_classe == "UKG":
                    fixed_path = path112u
                elif student_classe == "1":
                    fixed_path = path1121
                elif student_classe == "2":
                    fixed_path = path1122
                elif student_classe == "3":
                    fixed_path = path1123
                elif student_classe == "4":
                    fixed_path = path1124
                elif student_classe == "5":
                    fixed_path = path1125
                elif student_classe == "6":
                    fixed_path = path1126
                elif student_classe == "7":
                    fixed_path = path1127
                elif student_classe == "8":
                    fixed_path = path1128
                elif student_classe == "9":
                    fixed_path = path1129
                elif student_classe == "10":
                    fixed_path = path11210

                os.chdir(fixed_path)
                result_book = load_workbook(filename)
                result_sheet = result_book[filename[6:-5]]
                text1 = counter
                text2 = filename[0:5]
                text3 = result_sheet["C4"].value
                text4 = result_sheet["C5"].value
                text5 = result_sheet["C10"].value
                text6 = f'{result_sheet["C8"].value}-{result_sheet["D8"].value}-{result_sheet["E8"].value}'
                text7 = result_sheet["C13"].value

                def search_navi(event):
                    s_tree = search_tree.focus()
                    s_tree1 = search_tree.item(s_tree)
                    s_tree2 = s_tree1['values']
                    student_classe = s_tree2[3]

                    if student_classe == "Pre-Nursery":
                        fixed_path1 = path112pn
                    elif student_classe == "Nursery":
                        fixed_path1 = path112n
                    elif student_classe == "LKG":
                        fixed_path1 = path112l
                    elif student_classe == "UKG":
                        fixed_path1 = path112u
                    elif student_classe == 1:
                        fixed_path1 = path1121
                    elif student_classe == 2:
                        fixed_path1 = path1122
                    elif student_classe == 3:
                        fixed_path1 = path1123
                    elif student_classe == 4:
                        fixed_path1 = path1124
                    elif student_classe == 5:
                        fixed_path1 = path1125
                    elif student_classe == 6:
                        fixed_path1 = path1126
                    elif student_classe == 7:
                        fixed_path1 = path1127
                    elif student_classe == 8:
                        fixed_path1 = path1128
                    elif student_classe == 9:
                        fixed_path1 = path1129
                    elif student_classe == 10:
                        fixed_path1 = path11210

                    s_filename = s_tree2[1] + " " + s_tree2[2] + ".xlsx"
                    os.chdir(fixed_path1)
                    s_list = [fixed_path1, s_filename]
                    result_book = load_workbook(s_filename)
                    result_sheet = result_book[s_filename[6:-5]]
                    text1 = result_sheet["C4"].value
                    text2 = s_filename[0:5]
                    text3 = f'{result_sheet["C7"].value}/{result_sheet["D7"].value}/{result_sheet["E7"].value}'
                    text4 = result_sheet["C5"].value
                    text5 = result_sheet["E9"].value
                    text6 = result_sheet["C9"].value

                    existing_navigation2 = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
                    existing_navigation2.grid(row=0, column=0, sticky="nsew")

                    student_menu = Label(existing_navigation2, text="Student Menu", font=("verdana", 12), width=25,
                                         fg="black", bg="orange", pady=5, border=1, relief=GROOVE)
                    student_menu.grid(row=0, column=0, columnspan=5, sticky="ew")
                    student_menu1 = Label(existing_navigation2, text="Name", font=("verdana", 10),
                                          fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu1.grid(row=1, column=0, sticky="ew")
                    student_menu2 = Label(existing_navigation2, text="Code", font=("verdana", 10),
                                          fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu2.grid(row=2, column=0, sticky="ew")
                    student_menu3 = Label(existing_navigation2, text="Date of Birth", font=("verdana", 10),
                                          fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu3.grid(row=3, column=0, sticky="ew")
                    student_menu4 = Label(existing_navigation2, text="Class", font=("verdana", 10),
                                          fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu4.grid(row=4, column=0, sticky="ew")
                    student_menu5 = Label(existing_navigation2, text="Address(T)", font=("verdana", 10),
                                          fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu5.grid(row=5, column=0, sticky="ew")
                    student_menu6 = Label(existing_navigation2, text="District", font=("verdana", 10),
                                          fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu6.grid(row=6, column=0, sticky="ew")

                    student_menu11 = Label(existing_navigation2, text=text1, font=("verdana", 10),
                                           fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu11.grid(row=1, column=1, columnspan=4, sticky="ew")
                    student_menu21 = Label(existing_navigation2, text=text2, font=("verdana", 10),
                                           fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu21.grid(row=2, column=1, columnspan=4, sticky="ew")
                    student_menu31 = Label(existing_navigation2, text=text3, font=("verdana", 10),
                                           fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu31.grid(row=3, column=1, columnspan=4, sticky="ew")
                    student_menu41 = Label(existing_navigation2, text=text4, font=("verdana", 10),
                                           fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu41.grid(row=4, column=1, columnspan=4, sticky="ew")
                    student_menu51 = Label(existing_navigation2, text=text5, font=("verdana", 10),
                                           fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu51.grid(row=5, column=1, columnspan=4, sticky="ew")
                    student_menu61 = Label(existing_navigation2, text=text6, font=("verdana", 10), width=10,
                                           fg="black", pady=5, border=1, relief=GROOVE)
                    student_menu61.grid(row=6, column=1, columnspan=4, sticky="ew")

                    def details(event, opinion, list):
                        s_path = list[0]
                        s_filename = list[1]
                        os.chdir(s_path)
                        wb_details = load_workbook(s_filename)

                        details_frame = Frame(detail_pane_frame)
                        details_frame.grid(row=0, column=0, sticky="nsew")
                        details_title = Label(details_frame, bg="orange", fg="black", width=715)
                        details_title.grid(row=0, column=0, columnspan=10, sticky="ew")

                        details_frame1 = Frame(details_frame)
                        details_frame1.grid(row=1, column=0, columnspan=10, sticky="ew")

                        details_text = Text(details_frame1, width=715, font=("verdana", 10))
                        details_text.pack(side=TOP)

                        # TODO Define each of the following parameters and MAke the procedure of the execution in teh program

                        if opinion == "Status":
                            ws_details = wb_details[s_filename[6:-5]]

                            details_title.configure(text="Status")
                            details_text.insert("1.0", f'Current Status: \t {ws_details["C3"].value}')

                        if opinion == "Identity":
                            ws_details = wb_details[s_filename[6:-5]]
                            # TODO Errror due to None value in the data entry

                            age = date_converter_returning_age(ws_details["C7"].value, ws_details["D7"].value,
                                                               ws_details["E7"].value, year_today,
                                                               month_today, day_today)

                            details_title.configure(text="Personal Information")
                            details_text.insert("1.0", f"\n Hostel's Accomodation : \t {ws_details['C15'].value}")
                            details_text.insert("1.0", f'\n Previous School: \t {ws_details["C14"].value}')
                            details_text.insert("1.0", f"\n GrandFather's Name: \t {ws_details['C12'].value}")
                            details_text.insert("1.0", f"\n Mother's Name: \t {ws_details['C11'].value}")
                            details_text.insert("1.0", f"\n Father's Name: \t {ws_details['C10'].value}")
                            details_text.insert("1.0",
                                                f'\n Temporary Address: \t {ws_details["C9"].value}-{ws_details["D9"].value}-{ws_details["E9"].value}')
                            details_text.insert("1.0",
                                                f'\n Permanent Address: \t {ws_details["C8"].value}-{ws_details["D8"].value}-{ws_details["E8"].value}')
                            details_text.insert("1.0", f"\n Age: \t {age}")
                            details_text.insert("1.0",
                                                f'\n Date of Birth: \t {ws_details["C7"].value}/{ws_details["D7"].value}/{ws_details["E7"].value}')
                            details_text.insert("1.0", f'\n Class: \t {ws_details["C5"].value}')
                            details_text.insert("1.0", f'\n Gender: \t {ws_details["C6"].value}')
                            details_text.insert("1.0", f'Name: \t {ws_details["C4"].value}')

                    status = Button(existing_navigation2, text="Status", font=("verdana", 10), fg="black",
                                    bg="bisque", pady=5, border=1, relief=GROOVE)
                    status.grid(row=7, column=0, columnspan=5, sticky="ew")
                    status.bind("<Button-1>",
                                lambda event, opinion="Status", list=s_list: details(event, opinion, list))
                    identity = Button(existing_navigation2, text="Identity", font=("verdana", 10),
                                      fg="black",
                                      bg="bisque", pady=5, border=1, relief=GROOVE)
                    identity.grid(row=8, column=0, columnspan=5, sticky="ew")
                    identity.bind("<Button-1>",
                                  lambda event, opinion="Identity", list=s_list: details(event, opinion, list))
                    grades_scholarships = Button(existing_navigation2, text="Grades and Scholarships",
                                                 font=("verdana", 10), fg="black", bg="bisque", pady=5, border=1,
                                                 relief=GROOVE)
                    grades_scholarships.grid(row=9, column=0, columnspan=5, sticky="ew")
                    grades_scholarships.bind("<Button-1>",
                                             lambda event, opinion="Grades and Scholarships", list=s_list: details(
                                                 event, opinion, list))
                    fees_charges = Button(existing_navigation2, text="Fees and Charges", font=("verdana", 10),
                                          fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
                    fees_charges.grid(row=10, column=0, columnspan=5, sticky="ew")
                    fees_charges.bind("<Button-1>",
                                      lambda event, opinion="Fees and Charges", list=s_list: details(event, opinion,
                                                                                                     list))
                    remarks = Button(existing_navigation2, text="Remarks", font=("verdana", 10),
                                     fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
                    remarks.grid(row=11, column=0, columnspan=5, sticky="ew")
                    remarks.bind("<Button-1>",
                                 lambda event, opinion="Remarks", list=s_list: details(event, opinion, list))
                    graphical_records = Button(existing_navigation2, text="Graphical Records", font=("verdana", 10),
                                               fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
                    graphical_records.grid(row=12, column=0, columnspan=5, sticky="ew")
                    graphical_records.bind("<Button-1>",
                                           lambda event, opinion="Graphical Records", list=s_list: details(event,
                                                                                                           opinion,
                                                                                                           list))
                    hostel = Button(existing_navigation2, text="Hostel Provisions", font=("verdana", 10),
                                    fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
                    hostel.grid(row=13, column=0, columnspan=5, sticky="ew")
                    hostel.bind("<Button-1>",
                                lambda event, opinion="Hostel Provisions", list=s_list: details(event, opinion, list))

                slist = [fixed_path, filename]

                search_tree.insert("", index="end", iid=counter - 1,
                                   values=(text1, text2, text3, text4, text5, text6, text7))
                search_tree.bind("<Button-1>", search_navi)

            search_tree = ttk.Treeview(frame_searched)
            search_tree['columns'] = (
                "S.N.", "Code", "Name", "Class", "Father's Name", "Permanent Address", "Enrollment Year")
            search_tree.column("#0", width=0, minwidth=0, anchor="w")
            search_tree.column("S.N.", width=30, minwidth=10, anchor="e")
            search_tree.column("Code", width=60, minwidth=10, anchor="center")
            search_tree.column("Name", width=120, minwidth=10, anchor="center")
            search_tree.column("Class", width=40, minwidth=10, anchor="center")
            search_tree.column("Father's Name", width=140, minwidth=10, anchor="center")
            search_tree.column("Permanent Address", width=200, minwidth=10, anchor="center")
            search_tree.column("Enrollment Year", width=100, minwidth=10, anchor="center")
            search_tree.heading("#0", text="", anchor="center")
            search_tree.heading("S.N.", text="S.N.", anchor="center")
            search_tree.heading("Code", text="Code", anchor="center")
            search_tree.heading("Name", text="Name", anchor="center")
            search_tree.heading("Class", text="Class", anchor="center")
            search_tree.heading("Father's Name", text="Father's Name", anchor="center")
            search_tree.heading("Permanent Address", text="Permanent Address", anchor="center")
            search_tree.heading("Enrollment Year", text="Enrollment Year", anchor="center")
            search_tree.grid(row=0, column=0, sticky="nsew")

            if search_type.get() == "Search by Name":
                column = 5
                for i in range(6, search_sheet.max_row + 1):
                    alpha = search_sheet.cell(row=i, column=column).value
                    alpha1 = ""
                    alpha2 = ""
                    alpha3 = ""
                    alpha4 = ""
                    alpha5 = ""
                    alpha6 = ""
                    start = 0
                    end = len(alpha)
                    t = 1
                    for j in range(start,
                                   end):  # Todo Check for the length of the string check by removing the +1
                        if (alpha[j] != " " and t == 1):
                            alpha1 += alpha[j]
                        elif (alpha[j] != " " and t == 2):
                            alpha2 += alpha[j]
                        elif (alpha[j] != " " and t == 3):
                            alpha3 += alpha[j]
                        elif (alpha[j] != " " and t == 4):
                            alpha4 += alpha[j]
                        elif (alpha[j] != " " and t == 5):
                            alpha5 += alpha[j]
                        elif (alpha[j] != " " and t == 5):
                            alpha6 += alpha[j]
                        else:
                            t += 1

                    if ((((text == alpha1 or text == alpha2) or (text == alpha3 or text == alpha4)) or
                         ((text == alpha5 or text == alpha6) or text == f'{alpha1} {alpha2}')) or
                        ((text == f'{alpha1} {alpha3}' or text == f'{alpha1} {alpha4}') or
                         (text == f'{alpha1} {alpha5}' or text == f'{alpha1} {alpha6}'))) or \
                            (((text == f'{alpha2} {alpha3}' or text == f'{alpha1} {alpha2} {alpha3}') or
                              (text == f'{alpha1} {alpha3} {alpha4}' or text == f'{alpha1} {alpha4} {alpha5}')) or
                             ((
                                      text == f'{alpha1} {alpha2} {alpha3} {alpha4}' or text == f'{alpha1} {alpha3} {alpha4} {alpha5}') or
                              (text == f'{alpha1} {alpha2} {alpha3} {alpha4}' or
                               text == f'{alpha1} {alpha2} {alpha3} {alpha4} {alpha5} {alpha6}'))):
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
            elif search_type.get() == "Search by Class":
                column = 6
                for i in range(6, search_sheet.max_row + 1):
                    alpha = search_sheet.cell(row=i, column=column).value

                    if text == alpha:
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
            elif search_type.get() == "Search by Registration Year":
                column = 20
                for i in range(6, search_sheet.max_row + 1):
                    alpha = search_sheet.cell(row=i, column=column).value

                    if text == alpha:
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
            elif search_type.get() == "Search by Permanent Address":
                column1 = 11
                column2 = 12
                column3 = 13
                for i in range(6, search_sheet.max_row + 1):
                    alpha1 = search_sheet.cell(row=i, column=column1).value
                    alpha2 = search_sheet.cell(row=i, column=column2).value
                    alpha3 = search_sheet.cell(row=i, column=column3).value

                    if (text == alpha1 or text == alpha2) or text == alpha3:
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
            elif search_type.get() == "Search by Temporary Address":
                column1 = 14
                column2 = 15
                column3 = 16
                for i in range(6, search_sheet.max_row + 1):
                    alpha1 = search_sheet.cell(row=i, column=column1).value
                    alpha2 = search_sheet.cell(row=i, column=column2).value
                    alpha3 = search_sheet.cell(row=i, column=column3).value

                    if (text == alpha1 or text == alpha2) or text == alpha3:
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
            elif search_type.get() == "Search by Father's Name":
                column = 17
                for i in range(6, search_sheet.max_row + 1):
                    alpha = search_sheet.cell(row=i, column=column).value
                    alpha1 = ""
                    alpha2 = ""
                    alpha3 = ""
                    alpha4 = ""
                    alpha5 = ""
                    alpha6 = ""
                    start = 0
                    end = len(alpha)
                    t = 1
                    for j in range(start, end):  # Todo Check for the length of the string check by removing the +1
                        if (alpha[j] != " " and t == 1):
                            alpha1 += alpha[j]
                        elif (alpha[j] != " " and t == 2):
                            alpha2 += alpha[j]
                        elif (alpha[j] != " " and t == 3):
                            alpha3 += alpha[j]
                        elif (alpha[j] != " " and t == 4):
                            alpha4 += alpha[j]
                        elif (alpha[j] != " " and t == 5):
                            alpha5 += alpha[j]
                        elif (alpha[j] != " " and t == 5):
                            alpha6 += alpha[j]
                        else:
                            t += 1

                    if ((((text == alpha1 or text == alpha2) or (text == alpha3 or text == alpha4)) or
                         ((text == alpha5 or text == alpha6) or text == f'{alpha1} {alpha2}')) or
                        ((text == f'{alpha1} {alpha3}' or text == f'{alpha1} {alpha4}') or
                         (text == f'{alpha1} {alpha5}' or text == f'{alpha1} {alpha6}'))) or \
                            (((text == f'{alpha2} {alpha3}' or text == f'{alpha1} {alpha2} {alpha3}') or
                              (text == f'{alpha1} {alpha3} {alpha4}' or text == f'{alpha1} {alpha4} {alpha5}')) or
                             ((
                                      text == f'{alpha1} {alpha2} {alpha3} {alpha4}' or text == f'{alpha1} {alpha3} {alpha4} {alpha5}') or
                              (text == f'{alpha1} {alpha2} {alpha3} {alpha4}' or
                               text == f'{alpha1} {alpha2} {alpha3} {alpha4} {alpha5} {alpha6}'))):
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        print(stu_class, filename, counter)
                        button_creator(stu_class, filename, counter)
                        counter += 1
            else:
                columna1 = 5
                columna2 = 6
                columna3 = 20
                columna4 = 11
                columna5 = 12
                columna6 = 13
                columna7 = 14
                columna8 = 15
                columna9 = 16
                columna10 = 17
                for i in range(6, search_sheet.max_row + 1):

                    alphas1 = search_sheet.cell(row=i, column=columna1).value
                    alphas2 = search_sheet.cell(row=i, column=columna2).value
                    alphas3 = search_sheet.cell(row=i, column=columna3).value
                    alphas4 = search_sheet.cell(row=i, column=columna4).value
                    alphas5 = search_sheet.cell(row=i, column=columna5).value
                    alphas6 = search_sheet.cell(row=i, column=columna6).value
                    alphas7 = search_sheet.cell(row=i, column=columna7).value
                    alphas8 = search_sheet.cell(row=i, column=columna8).value
                    alphas9 = search_sheet.cell(row=i, column=columna9).value
                    alphas10 = search_sheet.cell(row=i, column=columna10).value

                    alpha1 = ""
                    alpha2 = ""
                    alpha3 = ""
                    alpha4 = ""
                    alpha5 = ""
                    alpha6 = ""
                    start = 0
                    end = len(alphas1)
                    t = 1
                    for j in range(start,
                                   end):  # Todo Check for the length of the string check by removing the +1
                        if (alphas1[j] != " " and t == 1):
                            alpha1 += alphas1[j]
                        elif (alphas1[j] != " " and t == 2):
                            alpha2 += alphas1[j]
                        elif (alphas1[j] != " " and t == 3):
                            alpha3 += alphas1[j]
                        elif (alphas1[j] != " " and t == 4):
                            alpha4 += alphas1[j]
                        elif (alphas1[j] != " " and t == 5):
                            alpha5 += alphas1[j]
                        elif (alphas1[j] != " " and t == 5):
                            alpha6 += alphas1[j]
                        else:
                            t += 1

                    if ((text == alpha1 or text == alpha2) or (text == alpha3 or text == alpha4)) or (
                            text == alpha5 or text == alpha6):
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
                    if text == alphas2:
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
                    if text == alphas3:
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
                    if (text == alphas4 or text == alphas5) or text == alphas6:
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
                    if (text == alphas7 or text == alphas8) or text == alphas9:
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1
                    alpha11 = ""
                    alpha12 = ""
                    alpha13 = ""
                    alpha14 = ""
                    alpha15 = ""
                    alpha16 = ""
                    start = 0
                    end = len(alphas10)
                    t = 1
                    for j in range(start,
                                   end):  # Todo Check for the length of the string check by removing the +1
                        if (alphas10[j] != " " and t == 1):
                            alpha11 += alphas10[j]
                        elif (alphas10[j] != " " and t == 2):
                            alpha12 += alphas10[j]
                        elif (alphas10[j] != " " and t == 3):
                            alpha13 += alphas10[j]
                        elif (alphas10[j] != " " and t == 4):
                            alpha14 += alphas10[j]
                        elif (alphas10[j] != " " and t == 5):
                            alpha15 += alphas10[j]
                        elif (alphas10[j] != " " and t == 5):
                            alpha16 += alphas10[j]
                        else:
                            t += 1

                    if ((text == alpha11 or text == alpha12) or (text == alpha13 or text == alpha14)) or (
                            text == alpha15 or text == alpha16):
                        stu_class = search_sheet.cell(row=i, column=6).value
                        code = search_sheet.cell(row=i, column=2).value
                        stu_name = search_sheet.cell(row=i, column=5).value
                        filename = f'{code} {stu_name}.xlsx'
                        button_creator(stu_class, filename, counter)
                        counter += 1

        existing_navigation1 = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
        existing_navigation1.grid(row=0, column=0, sticky="nsew")

        search_label = Label(existing_navigation1, text="Find a Student", font=("verdana", 12), width=24, fg="black",
                             bg="orange"
                                "", pady=5, border=1, relief=GROOVE)
        search_label.grid(row=0, column=0, columnspan=4, sticky="ew")

        search_entry_label_o = Label(existing_navigation1, text="Search Options", font=("verdana", 10), fg="black",
                                     bg="bisque", pady=5, border=1, relief=GROOVE)
        search_entry_label_o.grid(row=1, column=0, columnspan=4, sticky="ew")

        search_label = ttk.Combobox(existing_navigation1, state="readonly", values=options_search,
                                    font=("verdana", 10), textvariable=search_type)
        search_label.grid(row=2, column=0, columnspan=4, sticky="ew")
        search_entry_label = Label(existing_navigation1, text="Enter Here:", font=("verdana", 10), fg="black",
                                   bg="bisque", pady=5, border=1, relief=GROOVE)
        search_entry_label.grid(row=3, column=0, sticky="w")
        search_entry = Entry(existing_navigation1, font=("verdana", 10), textvariable=search_text)
        search_entry.grid(row=3, column=1, columnspan=3, sticky="ew")

        search_enter = Button(existing_navigation1, text="Search", font=("verdana", 10), fg="black",
                              bg="bisque", pady=5, border=1, relief=GROOVE)
        search_enter.grid(row=4, column=0, columnspan=4, sticky="ew")
        search_enter.bind("<Button-1>", search_student)

    def academic_input(event):
        academic_navigation = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
        academic_navigation.grid(row=0, column=0, sticky="nsew")

        def program_input(event):
            pass

        existing_record = Button(academic_navigation, text="Program Input", font=("verdana", 12), width=24,
                                 fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
        existing_record.bind("<Button-1>", program_input)
        existing_record.grid(row=0, column=0, sticky="ew")

        def excel_input(event):
            g_class = StringVar()
            g_year = StringVar()

            excel_1 = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
            excel_1.grid(row=0, column=0, sticky="nsew")

            def generate_sheets(event):
                options_years = (
                    "2070", "2071", "2072", "2073", "2074", "2075", "2076", "2077", "2078", "2079", "2080", "2081",
                    "2082",
                    "2083",
                    "2084", "2085", "2086", "2087", "2088")
                options_classes = (
                    "Pre-Nursery", "Nursery", "LKG", "UKG", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10",
                    "All Class")

                def generator(event):
                    year = g_year.get()

                    def generator2(year, name_c):
                        g_class1 = name_c
                        filename1 = "Courses for Year " + year
                        filename2 = f'Class {g_class1}.xlsx'
                        path_g = path_reg + "\\" + filename1
                        path_desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
                        os.chdir(path_desktop)
                        folder_results = "Result Sheets"
                        path_sheets = os.path.join(path_desktop, folder_results)
                        try:
                            os.chdir(path_sheets)
                        except:
                            os.makedirs(folder_results)

                        w_book = Workbook()
                        w_sheet = w_book.active
                        w_sheet.title = filename2[0:-5]
                        try:
                            os.chdir(path_g)
                            try:
                                g_book = load_workbook(filename2)
                                g_sheet = g_book[filename2[0:-5]]
                                w_sheet.cell(row=5, column=1).value = "S.N."
                                w_sheet.cell(row=5, column=2).value = "Student Code"
                                w_sheet.cell(row=5, column=3).value = "Student Name"
                                index = 4
                                for i in range(1, 1 + g_sheet.max_row):
                                    if g_sheet.cell(row=i, column=1).value != None:
                                        w_sheet.cell(row=5, column=index).value = g_sheet.cell(row=i, column=2).value
                                        index += 1
                                        if g_sheet.cell(row=i, column=3).value != "Theory and Practical":
                                            if g_sheet.cell(row=i, column=3).value == "Practical":
                                                w_sheet.cell(row=6, column=index - 1).value = "Practical"
                                            else:
                                                w_sheet.cell(row=6, column=index - 1).value = "Theory"

                                        else:
                                            w_sheet.cell(row=6, column=index - 1).value = "Theory"
                                            w_sheet.cell(row=6, column=index).value = "Practical"
                                            index += 1
                                    else:
                                        pass

                                counter = 1
                                column = 6
                                os.chdir(path114)
                                search_book = load_workbook("All Students.xlsx")
                                search_sheet = search_book["All Students"]
                                for i in range(6, search_sheet.max_row + 1):
                                    alpha = search_sheet.cell(row=i, column=column).value

                                    if g_class1 == alpha:
                                        code = search_sheet.cell(row=i, column=2).value
                                        stu_name = search_sheet.cell(row=i, column=5).value
                                        w_sheet.cell(row=counter + 6, column=1).value = counter
                                        w_sheet.cell(row=counter + 6, column=2).value = code
                                        w_sheet.cell(row=counter + 6, column=3).value = stu_name
                                        counter += 1

                                os.chdir(path_sheets)
                                w_book.save(filename2)
                            except:
                                dialog = Tk()
                                labela = Label(dialog, text=f"Courses not registered for Class {g_class1}!")
                                labela.grid(row=0, column=0, columnspan=5)
                        except:
                            dialog = Tk()
                            labela = Label(dialog, text=f"File not found for Year {year}!")
                            labela.grid(row=0, column=0, columnspan=5)

                    if g_class.get() != "All Class":
                        generator2(year, g_class.get())

                    else:
                        for item in options_classes:
                            if item != "All Class":
                                generator2(year, item)
                            else:
                                pass

                excel_2 = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
                excel_2.grid(row=0, column=0, sticky="nsew")

                existing_record = Label(excel_2, text="Select Year", font=("verdana", 10), width=30,
                                        fg="black", border=1, relief=GROOVE)
                existing_record.grid(row=0, column=0, sticky="ew")

                g_year1 = ttk.Combobox(excel_2, state="readonly", values=options_years,
                                       font=("verdana", 10), textvariable=g_year)
                g_year1.grid(row=1, column=0, columnspan=4, sticky="ew")
                g_year1.set("2078")

                existing_record = Label(excel_2, text="Select Class", font=("verdana", 10),
                                        fg="black", border=1, relief=GROOVE)
                existing_record.grid(row=2, column=0, sticky="ew")

                g_year1 = ttk.Combobox(excel_2, state="readonly", values=options_classes,
                                       font=("verdana", 10), textvariable=g_class)
                g_year1.grid(row=3, column=0, columnspan=4, sticky="ew")

                existing_record1 = Button(excel_2, text="Submit", font=("verdana", 12),
                                          fg="black", bg="springgreen2", border=1, relief=GROOVE)
                existing_record1.grid(row=4, column=0, sticky="ew")

                existing_record1.bind("<Button-1>", generator)

            existing_record = Button(excel_1, text="Generate Excel Sheets", font=("verdana", 12), width=24,
                                     fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", generate_sheets)
            existing_record.grid(row=0, column=0, sticky="ew")

            # TODO after confirming the subjects of the class from the pre nursery section to the class10
            def load_sheets(event):

                excel_3 = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
                excel_3.grid(row=0, column=0, sticky="nsew")

                existing_record = Button(excel_3, text="Provide the Appropriate Path", font=("verdana", 12), width=24,
                                         fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
                existing_record.grid(row=0, column=0, sticky="ew")

            existing_record = Button(excel_1, text="Load Excel Sheet", font=("verdana", 12), width=24,
                                     fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", load_sheets)
            existing_record.grid(row=1, column=0, sticky="ew")

        existing_record = Button(academic_navigation, text="Excel Sheet Input", font=("verdana", 12), width=24,
                                 fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
        existing_record.bind("<Button-1>", excel_input)
        existing_record.grid(row=1, column=0, sticky="ew")

        def search_student(event):
            search_type = StringVar()
            search_text = StringVar()

            existing_navigation = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
            existing_navigation.grid(row=0, column=0, sticky="nsew")

            options_search = (
                "", "Search by Name", "Search by Class", "Search by Registration Year", "Search by Permanent Address",
                "Search by Temporary Address", "Search by Father's Name")

            def search_student(event):
                pass

            existing_navigation1 = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
            existing_navigation1.grid(row=0, column=0, sticky="nsew")

            search_label = Label(existing_navigation1, text="Find a Student", font=("verdana", 12), width=24,
                                 fg="black",
                                 bg="orange"
                                    "", pady=5, border=1, relief=GROOVE)
            search_label.grid(row=0, column=0, columnspan=4, sticky="ew")

            search_entry_label_o = Label(existing_navigation1, text="Search Options", font=("verdana", 10),
                                         fg="black",
                                         bg="bisque", pady=5, border=1, relief=GROOVE)
            search_entry_label_o.grid(row=1, column=0, columnspan=4, sticky="ew")

            search_label = ttk.Combobox(existing_navigation1, state="readonly", values=options_search,
                                        font=("verdana", 10), textvariable=search_type)
            search_label.grid(row=2, column=0, columnspan=4, sticky="ew")
            search_entry_label = Label(existing_navigation1, text="Enter Here:", font=("verdana", 10), fg="black",
                                       bg="bisque", pady=5, border=1, relief=GROOVE)
            search_entry_label.grid(row=3, column=0, sticky="w")
            search_entry = Entry(existing_navigation1, font=("verdana", 10), textvariable=search_text)
            search_entry.grid(row=3, column=1, columnspan=3, sticky="ew")

            search_enter = Button(existing_navigation1, text="Search", font=("verdana", 10), fg="black",
                                  bg="bisque", pady=5, border=1, relief=GROOVE)
            search_enter.grid(row=4, column=0, columnspan=4, sticky="ew")
            search_enter.bind("<Button-1>", search_student)

        existing_record = Button(academic_navigation, text="Data Correction (Admin)", font=("verdana", 12), width=24,
                                 fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
        existing_record.bind("<Button-1>", search_student)
        existing_record.grid(row=2, column=0, sticky="ew")

        def courses_registration(event):
            courses = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
            courses.grid(row=0, column=0, sticky="nsew")

            def register_data(event, c_name):
                options_courses = (
                    "Nepali", "English", "Mathematics", "Science", "Social Studies", "General Knowledge",
                    "Population Studies", "Health Education", "Environment Science", "Grammar", "Byakaran", "Account",
                    "Business Studies",
                    "Health Population and Environment",
                    "Computer Science", "Optional Mathematics", "Occupation",
                    "Physical Training", "Games", "Music", "Drawing", "Literature")
                options_years = (
                    "2070", "2071", "2072", "2073", "2074", "2075", "2076", "2077", "2078", "2079", "2080", "2081",
                    "2082",
                    "2083",
                    "2084", "2085", "2086", "2087", "2088")
                courses_options = ("Theory", "Practical", "Theory and Practical")
                courses_e0 = Frame(detail_pane_frame, relief=FLAT)
                courses_e0.grid(row=0, column=0, sticky="nsew")
                courses_e1 = Frame(courses_e0, relief=FLAT)
                courses_e1.grid(row=0, column=0, sticky="nsew")
                courses_e = Frame(courses_e0, relief=FLAT)
                courses_e.grid(row=1, column=0, sticky="nsew")
                courses_e2 = Frame(courses_e0, relief=FLAT)
                courses_e2.grid(row=2, column=0, sticky="nsew")

                label1 = Label(courses_e1, text=f'Class-{c_name} Registration Form', font=("verdana", 12), width=70,
                               bg="coral", pady=10)
                label1.grid(row=0, column=0, columnspan=4, sticky="ew")

                labels = Label(courses_e1, text=f'Registration For Year', font=("verdana", 10, "bold"), pady=5)
                labels.grid(row=4, column=0)
                entrysy = ttk.Combobox(courses_e1, values=options_years, state="readonly", font=("verdana", 10, "bold"),
                                       width=30)
                entrysy.grid(row=4, column=1, columnspan=2)
                entrysy.set("2078")

                label2 = Label(courses_e, text=f'Core Subjects', font=("verdana", 10, "bold"), pady=10)
                label2.grid(row=1, column=0)

                labels = Label(courses_e, text=f'First Subject', font=("verdana", 10), pady=5)
                labels.grid(row=2, column=0)
                entrys1 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys1.grid(row=2, column=1, columnspan=1)
                entrys11 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys11.grid(row=2, column=2, columnspan=1)
                entrys11.set("Theory")

                labels = Label(courses_e, text=f'Second Subject', font=("verdana", 10), pady=5)
                labels.grid(row=3, column=0)
                entrys2 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys2.grid(row=3, column=1, columnspan=1)
                entrys21 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys21.grid(row=3, column=2, columnspan=1)
                entrys21.set("Theory")

                labels = Label(courses_e, text=f'Third Subject', font=("verdana", 10), pady=5)
                labels.grid(row=4, column=0)
                entrys3 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys3.grid(row=4, column=1, columnspan=1)
                entrys31 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys31.grid(row=4, column=2, columnspan=1)
                entrys31.set("Theory")

                labels = Label(courses_e, text=f'Forth Subject', font=("verdana", 10), pady=5)
                labels.grid(row=5, column=0)
                entrys4 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys4.grid(row=5, column=1, columnspan=1)
                entrys41 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys41.grid(row=5, column=2, columnspan=1)
                entrys41.set("Theory")

                labels = Label(courses_e, text=f'Fifth Subject', font=("verdana", 10), pady=5)
                labels.grid(row=6, column=0)
                entrys5 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys5.grid(row=6, column=1, columnspan=1)
                entrys51 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys51.grid(row=6, column=2, columnspan=1)
                entrys51.set("Theory")

                labels = Label(courses_e, text=f'Sixth Subject', font=("verdana", 10), pady=5)
                labels.grid(row=7, column=0)
                entrys6 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys6.grid(row=7, column=1, columnspan=1)
                entrys61 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys61.grid(row=7, column=2, columnspan=1)
                entrys61.set("Theory")

                labels = Label(courses_e, text=f'Seventh Subject', font=("verdana", 10), pady=5)
                labels.grid(row=8, column=0)
                entrys7 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys7.grid(row=8, column=1, columnspan=1)
                entrys71 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys71.grid(row=8, column=2, columnspan=1)

                labels = Label(courses_e, text=f'Eighth Subject', font=("verdana", 10), pady=5)
                labels.grid(row=9, column=0)
                entrys8 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys8.grid(row=9, column=1, columnspan=1)
                entrys81 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys81.grid(row=9, column=2, columnspan=1)

                labels = Label(courses_e, text=f'Ninth Subject', font=("verdana", 10), pady=5)
                labels.grid(row=10, column=0)
                entrys9 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                       width=30)
                entrys9.grid(row=10, column=1, columnspan=1)
                entrys91 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                        width=20)
                entrys91.grid(row=10, column=2, columnspan=1)

                labels = Label(courses_e, text=f'Tenth Subject', font=("verdana", 10), pady=5)
                labels.grid(row=11, column=0)
                entrys10 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                        width=30)
                entrys10.grid(row=11, column=1, columnspan=1)
                entrys101 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                         width=20)
                entrys101.grid(row=11, column=2, columnspan=1)

                labels = Label(courses_e, text=f'Eleventh Subject', font=("verdana", 10), pady=5)
                labels.grid(row=12, column=0)
                entrys11 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                        width=30)
                entrys11.grid(row=12, column=1, columnspan=1)
                entrys111 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                         width=20)
                entrys111.grid(row=12, column=2, columnspan=1)

                labels = Label(courses_e, text=f'Twelveth Subject', font=("verdana", 10), pady=5)
                labels.grid(row=13, column=0)
                entrys12 = ttk.Combobox(courses_e, values=options_courses, state="readonly", font=("verdana", 10),
                                        width=30)
                entrys12.grid(row=13, column=1, columnspan=1)
                entrys121 = ttk.Combobox(courses_e, values=courses_options, state="readonly", font=("verdana", 10),
                                         width=20)
                entrys121.grid(row=13, column=2, columnspan=1)

                label2 = Label(courses_e2, text=f'Optional Subject 1', font=("verdana", 10, "bold"), pady=10)
                label2.grid(row=14, column=0)
                entrys131 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=19)
                entrys131.grid(row=15, column=0)
                entrys132 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=19)
                entrys132.grid(row=15, column=1)
                entrys133 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=18)
                entrys133.grid(row=15, column=2)
                entrys134 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=17)
                entrys134.grid(row=15, column=3)
                entrys1311 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=19)
                entrys1311.grid(row=16, column=0)
                entrys1322 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=19)
                entrys1322.grid(row=16, column=1)
                entrys1333 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=18)
                entrys1333.grid(row=16, column=2)
                entrys1344 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=17)
                entrys1344.grid(row=16, column=3)

                label2 = Label(courses_e2, text=f'Optional Subject 2', font=("verdana", 10, "bold"), pady=10)
                label2.grid(row=17, column=0)
                entrys141 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=19)
                entrys141.grid(row=18, column=0)
                entrys142 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=19)
                entrys142.grid(row=18, column=1)
                entrys143 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=18)
                entrys143.grid(row=18, column=2)
                entrys144 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=17)
                entrys144.grid(row=18, column=3)
                entrys1411 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=19)
                entrys1411.grid(row=19, column=0)
                entrys1422 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=19)
                entrys1422.grid(row=19, column=1)
                entrys1433 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=18)
                entrys1433.grid(row=19, column=2)
                entrys1444 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=17)
                entrys1444.grid(row=19, column=3)
                label2 = Label(courses_e2, text=f'Optional Subject 3', font=("verdana", 10, "bold"), pady=10)
                label2.grid(row=20, column=0)
                entrys151 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=19)
                entrys151.grid(row=21, column=0)
                entrys152 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=19)
                entrys152.grid(row=21, column=1)
                entrys153 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=18)
                entrys153.grid(row=21, column=2)
                entrys154 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=17)
                entrys154.grid(row=21, column=3)
                entrys1511 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=19)
                entrys1511.grid(row=22, column=0)
                entrys1522 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=19)
                entrys1522.grid(row=22, column=1)
                entrys1533 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=18)
                entrys1533.grid(row=22, column=2)
                entrys1544 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=17)
                entrys1544.grid(row=22, column=3)

                label2 = Label(courses_e2, text=f'Optional Subject 4', font=("verdana", 10, "bold"), pady=10)
                label2.grid(row=23, column=0)
                entrys161 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=19)
                entrys161.grid(row=24, column=0)
                entrys162 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=19)
                entrys162.grid(row=24, column=1)
                entrys163 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=18)
                entrys163.grid(row=24, column=2)
                entrys164 = ttk.Combobox(courses_e2, values=options_courses, state="readonly", font=("verdana", 10),
                                         width=17)
                entrys164.grid(row=24, column=3)
                entrys1611 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=19)
                entrys1611.grid(row=25, column=0)
                entrys1622 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=19)
                entrys1622.grid(row=25, column=1)
                entrys1633 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=18)
                entrys1633.grid(row=25, column=2)
                entrys1644 = ttk.Combobox(courses_e2, values=courses_options, state="readonly", font=("verdana", 10),
                                          width=17)
                entrys1644.grid(row=25, column=3)

                text1 = f'Note \n -Any number of Subjects can be entered. \n -Subjects should be changed every year if courses varies. \n -At least two subjects should be available for Optional. \n -If not needed Optional Subjects may be skipped.'
                texts = Text(courses_e2, font=("verdana", 10), width=75, height=6)
                texts.grid(row=26, column=0, columnspan=4)
                texts.insert('1.0', text1)
                texts.configure(state="disabled")

                def confirm(event, reg_list):
                    date_time = date_time1()
                    folder_name = "Courses for Year " + reg_list[1]
                    path_r = path_reg + "\\" + folder_name
                    try:
                        os.chdir(path_r)
                    except:
                        os.chdir(path_reg)
                        os.makedirs(folder_name)
                        os.chdir(path_r)

                    filename = f'Class {reg_list[0]}.xlsx'
                    '''

                    if os.path.isfile(filename) == TRUE:
                        overwrite = Tk()
                        labela = Label(overwrite, text="Do you want to overwrite in previously written Data?")
                        labela.grid(row=0, column=0, columnspan=5)

                        def yes(event):
                            w_reg = load_workbook(filename) # TODO w_reg referenced before assignment error and the window closing procedure shuld be established

                        def no(event):
                            overwrite.quit() # TODO It closes the whole program but only the opoup window needs to be closed

                        c_button = Button(overwrite, text="Yes", width=20)
                        c_button.grid(row=1, column=0, columnspan=2)
                        c_button.bind("<Button-1>", yes)
                        c_button = Button(overwrite, text="No", width=20)
                        c_button.grid(row=1, column=3, columnspan=2)
                        c_button.bind("<Button-1>", no)

                    elsee: '''
                    w_reg = Workbook()

                    s1 = entrys1.get()
                    s2 = entrys2.get()
                    s3 = entrys3.get()
                    s4 = entrys4.get()
                    s5 = entrys5.get()
                    s6 = entrys6.get()
                    s7 = entrys7.get()
                    s8 = entrys8.get()
                    s9 = entrys9.get()
                    s10 = entrys10.get()
                    s11 = entrys11.get()
                    s12 = entrys12.get()
                    s131 = entrys131.get()
                    s132 = entrys132.get()
                    s133 = entrys133.get()
                    s134 = entrys134.get()
                    s141 = entrys141.get()
                    s142 = entrys142.get()
                    s143 = entrys143.get()
                    s144 = entrys144.get()
                    s151 = entrys151.get()
                    s152 = entrys152.get()
                    s153 = entrys153.get()
                    s154 = entrys154.get()
                    s161 = entrys161.get()
                    s162 = entrys162.get()
                    s163 = entrys163.get()
                    s164 = entrys164.get()

                    s11 = entrys11.get()
                    s21 = entrys21.get()
                    s31 = entrys31.get()
                    s41 = entrys41.get()
                    s51 = entrys51.get()
                    s61 = entrys61.get()
                    s71 = entrys71.get()
                    s81 = entrys81.get()
                    s91 = entrys9.get()
                    s101 = entrys101.get()
                    s111 = entrys111.get()
                    s121 = entrys121.get()
                    s1311 = entrys1311.get()
                    s1322 = entrys1322.get()
                    s1333 = entrys1333.get()
                    s1344 = entrys1344.get()
                    s1411 = entrys1411.get()
                    s1422 = entrys1422.get()
                    s1433 = entrys1433.get()
                    s1444 = entrys1444.get()
                    s1511 = entrys1511.get()
                    s1522 = entrys1522.get()
                    s1533 = entrys1533.get()
                    s1544 = entrys1544.get()
                    s1611 = entrys1611.get()
                    s1622 = entrys1622.get()
                    s1633 = entrys1633.get()
                    s1644 = entrys1644.get()

                    core_subjects = (s1, s2, s3, s4, s5, s6, s7, s8, s9, s10, s11, s12)
                    optional_subject1 = (s131, s132, s133, s134)
                    optional_subject2 = (s141, s142, s143, s144)
                    optional_subject3 = (s151, s152, s153, s154)
                    optional_subject4 = (s161, s162, s163, s164)

                    ocore_subjects = (s11, s21, s31, s41, s51, s61, s71, s81, s91, s101, s111, s121)
                    ooptional_subject1 = (s1311, s1322, s1333, s1344)
                    ooptional_subject2 = (s1411, s1422, s1433, s1444)
                    ooptional_subject3 = (s1511, s1522, s1533, s1544)
                    ooptional_subject4 = (s1611, s1622, s1633, s1644)

                    ws_reg = w_reg.active
                    ws_reg.title = filename[0:-5]

                    start0 = 1
                    for i in range(12):
                        if core_subjects[i] != "":
                            ws_reg.cell(row=start0 + 4, column=1).value = start0
                            ws_reg.cell(row=start0 + 4, column=2).value = core_subjects[i]
                            ws_reg.cell(row=start0 + 4, column=3).value = ocore_subjects[i]
                            start0 += 1
                        else:
                            pass

                    if s131 != "":
                        start1 = 5
                        ws_reg.cell(row=start0 + 5, column=2).value = "Optional Coarses 1"
                        ws_reg.cell(row=start0 + 5, column=3).value = "Type"
                        for i in range(4):
                            if optional_subject1[i] != "":
                                ws_reg.cell(row=start0 + 6, column=1).value = start0
                                ws_reg.cell(row=start0 + 6, column=2).value = optional_subject1[i]
                                ws_reg.cell(row=start0 + 6, column=3).value = ooptional_subject1[i]
                                start1 += 1
                                start0 += 1
                            else:
                                pass
                        if s141 != "":
                            start2 = 5
                            ws_reg.cell(row=start0 + 7, column=2).value = "Optional Coarses 2"
                            ws_reg.cell(row=start0 + 7, column=3).value = "Type"
                            for i in range(4):
                                if optional_subject2[i] != "":
                                    ws_reg.cell(row=start0 + 8, column=1).value = start0
                                    ws_reg.cell(row=start0 + 8, column=2).value = optional_subject2[i]
                                    ws_reg.cell(row=start0 + 8, column=3).value = ooptional_subject2[i]
                                    start2 += 1
                                    start0 += 1
                                else:
                                    pass
                            if s151 != "":
                                start3 = 5
                                ws_reg.cell(row=start0 + 9, column=2).value = "Optional Coarses 3"
                                ws_reg.cell(row=start0 + 9, column=3).value = "Type"
                                for i in range(4):
                                    if optional_subject3[i] != "":
                                        ws_reg.cell(row=start0 + 10, column=1).value = start0
                                        ws_reg.cell(row=start0 + 10, column=2).value = optional_subject3[i]
                                        ws_reg.cell(row=start0 + 10, column=3).value = ooptional_subject3[i]
                                        start3 += 1
                                        start0 += 1
                                    else:
                                        pass
                                if s161 != "":
                                    start4 = 5
                                    ws_reg.cell(row=start0 + 11, column=2).value = "Optional Coarses 4"
                                    ws_reg.cell(row=start0 + 11, column=3).value = "Type"
                                    for i in range(4):
                                        if optional_subject4[i] != "":
                                            ws_reg.cell(row=start0 + 12, column=1).value = start0
                                            ws_reg.cell(row=start0 + 12, column=2).value = optional_subject4[i]
                                            ws_reg.cell(row=start0 + 12, column=3).value = ooptional_subject4[i]
                                            start4 += 1
                                            start0 += 1
                                        else:
                                            pass
                                else:
                                    pass
                            else:
                                pass
                        else:
                            pass
                    else:
                        pass
                    ws_reg.cell(row=3, column=2).value = "Registered Coarses"
                    ws_reg.cell(row=4, column=2).value = "Core Coarses"
                    ws_reg.cell(row=4, column=3).value = "Type"
                    ws_reg.cell(row=2, column=2).value = "Updated On:"
                    ws_reg.cell(row=2, column=3).value = date_time[0:10]

                    w_reg.save(filename)
                    text = f"{filename[0:-5]} Course has successfuly been registered in {date_time}\n"
                    act_text.configure(state='normal')
                    act_text.insert('1.0', f'\n {text}')
                    act_text.configure(state='disabled')

                s_confirm = Button(courses_e2, text="Confirm", font=("verdana", 12), bg="springgreen2", width=30)
                s_confirm.grid(row=27, column=1, columnspan=2)
                s_confirm.bind("<Button-1>", lambda event, reg_list=[c_name, entrysy.get()]: confirm(event, reg_list))

            existing_record = Button(courses, text="Pre-Nursery", font=("verdana", 10), width=30,
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="Pre-Nursery": register_data(event, c_name))
            existing_record.grid(row=0, column=0, sticky="ew")

            existing_record = Button(courses, text="Nursery", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="Nursery": register_data(event, c_name))
            existing_record.grid(row=1, column=0, sticky="ew")

            existing_record = Button(courses, text="LKG", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="LKG": register_data(event, c_name))
            existing_record.grid(row=2, column=0, sticky="ew")

            existing_record = Button(courses, text="UKG", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="UKG": register_data(event, c_name))
            existing_record.grid(row=3, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-1", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="1": register_data(event, c_name))
            existing_record.grid(row=4, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-2", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="2": register_data(event, c_name))
            existing_record.grid(row=5, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-3", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="3": register_data(event, c_name))
            existing_record.grid(row=6, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-4", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="4": register_data(event, c_name))
            existing_record.grid(row=7, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-5", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="5": register_data(event, c_name))
            existing_record.grid(row=8, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-6", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="6": register_data(event, c_name))
            existing_record.grid(row=9, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-7", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="7": register_data(event, c_name))
            existing_record.grid(row=10, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-8", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="8": register_data(event, c_name))
            existing_record.grid(row=11, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-9", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="9": register_data(event, c_name))
            existing_record.grid(row=12, column=0, sticky="ew")

            existing_record = Button(courses, text="Class-10", font=("verdana", 10),
                                     fg="black", bg="bisque", border=1, relief=GROOVE)
            existing_record.bind("<Button-1>", lambda event, c_name="10": register_data(event, c_name))
            existing_record.grid(row=13, column=0, sticky="ew")

        existing_record = Button(academic_navigation, text="Courses Registration", font=("verdana", 12), width=24,
                                 fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
        existing_record.bind("<Button-1>", courses_registration)
        existing_record.grid(row=3, column=0, sticky="ew")

    def students_event(event):
        students_navigation = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
        students_navigation.grid(row=0, column=0, sticky="nsew")
        # students_navigation.rowconfigure(0, weight=1)
        # students_navigation.columnconfigure(0, weight=1)

        new_registration = Button(students_navigation, text="Registration", font=("verdana", 12), width=24,
                                  relief=GROOVE,
                                  fg="black", bg="bisque", pady=5, border=1)
        new_registration.bind("<Button-1>", new_registration_event)
        new_registration.grid(row=0, column=0, sticky="ew")
        existing_record = Button(students_navigation, text="Existing Record", font=("verdana", 12), width=24,
                                 fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
        existing_record.bind("<Button-1>", existing_record_event)
        existing_record.grid(row=1, column=0, sticky="ew")

        existing_record = Button(students_navigation, text="Academic Performance", font=("verdana", 12), width=24,
                                 fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
        existing_record.bind("<Button-1>", academic_input)
        existing_record.grid(row=2, column=0, sticky="ew")

    def new_registration_event1(event):
        def teacher_to_db(event):
            date_time = date_time1()

            t1 = teacher_name.get()
            t2 = teacher_age.get()
            t3 = teacher_gender.get()
            t4 = teacher_dob1.get()
            t4a = teacher_dob2.get()
            t4b = teacher_dob3.get()
            t5 = teacher_marriage.get()
            t6 = teacher_paddress.get()
            t6a = teacher_paddress1.get()
            t6b = teacher_paddress2.get()
            t7 = teacher_taddress.get()
            t7a = teacher_taddress1.get()
            t7b = teacher_taddress2.get()
            t8 = teacher_contact1.get()
            t9 = teacher_contact2.get()
            t10 = teacher_email.get()
            t11 = teacher_facebook.get()
            # teacher_photo = StringVar()
            t12 = teacher_level1.get()
            t13 = teacher_board1.get()
            t14 = teacher_institution1.get()
            t15 = teacher_grade1.get()
            t15a = teacher_subject.get()
            t16 = teacher_fow1.get()
            t17 = teacher_post1.get()
            t18 = teacher_duration1.get()
            t19 = teacher_wintitution1.get()
            t20 = teacher_wlocation1.get()
            t21 = teacher_description_entry.get('1.0', 'end-1c')
            t22 = teacher_mainsubject.get()
            t23 = teacher_class1.get()
            t23a = teacher_class2.get()
            t24 = teacher_ssubject1.get()
            t25 = teacher_class21.get()
            t25a = teacher_class22.get()
            t26 = teacher_ssubject2.get()
            t27 = teacher_class31.get()
            t27a = teacher_class32.get()

            composite_title = ["S.N.", "Name of the Teacher", "Age", "Gender", "Date of Birth", " ", " ",
                               "Marital Status", "Permanent Address", " ", " ", "Temporary Address", " ", " ",
                               "Contact Number",
                               "Secondary Contact Number",
                               "E-mail Address", "Facebook ID", "Academic Qualification", " ", " ", " ", " ",
                               "Experience and Eligibility",
                               " ", " ", " ", " ", "Description of the teacher", "Courses of Teaching", " ", " ", " ",
                               " ", " "]
            location = ["District", "Local Government", "Area/Tol"]
            dob_title = ["Year", "Month", "Day"]
            education_title = ["Education Level",
                               "Education Board", "Institution of Study", "Percentage/Grade", "Main Subject"]
            experience_title = ["Field of Work", "Working Post", "Duration", "Institute", "Location of Work"]
            coarses_title = ["Main Teaching Subject",
                             "Class of Teaching", "Secondary Teaching Subject - 1", "Class of Teaching",
                             "Secondary Teaching Subject - 2", "Class of Teaching"]
            if t23a != "":
                crange1 = t23 + " to " + t23a
            else:
                crange1 = t23
            if t25a != "":
                crange2 = t25 + " to " + t25a
            else:
                crange2 = t25
            if t27a != "":
                crange3 = t27 + " to " + t27a
            else:
                crange3 = t27
            composite_data = [t1, t2, t3, t4, t4a, t4b, t5, t6, t6a, t6b, t7, t7a, t7b, t8, t9, t10, t11, t12, t13,
                              t14, t15, t15a, t16, t17, t18, t19, t20, t21, t22,
                              crange1, t24, crange2, t26, crange3]

            def individual_writer():
                os.chdir(path122)

                def write_class(cell_value, cell_no, extra, numb):
                    if cell_value != "":
                        c_v = "D" + str(cell_no)
                        worksheet1[c_v].value = extra
                        if numb == 2:
                            c_v = "E" + str(cell_no)
                            worksheet1[c_v].value = cell_value

                filename = teacher_name.get() + str(teacher_age.get())
                workbook1 = Workbook()
                worksheet1 = workbook1.active
                worksheet1.title = teacher_name.get()
                i = 4
                j = 5
                data_title1 = ["Personal Details", "Name of the Teacher", "Age", "Gender", "Date of Birth",
                               "Marital Status",
                               "Permanent Address", "Temporary Address", "Contact Number",
                               "Secondary Contact Number",
                               "E-mail Address", "Facebook ID", "Academic Qualification", "Education Level",
                               "Education Board", "Institution of Study", "Percentage/Grade", "Main Subject",
                               "Experience and Eligibility",
                               "Field of Work", "Working Post", "Duration", "Institute", "Location of Work",
                               "Description of the teacher", "Courses of Teaching", "Main Teaching Subject",
                               "Class of Teaching", "Secondary Teaching Subject - 1", "Class of Teaching",
                               "Secondary Teaching Subject - 2",
                               "Class of Teaching"]
                entry_datas1 = [t1, t2, t3, t4, t5, t6, t7, t8, t9, t10, t11, t12, t13, t14, t15, t15a, t16, t17,
                                "",
                                t19, t20, t21, t22, t23, t24, t25, t26, t27]
                if t16 != "":
                    worksheet1["B27"] = t18
                for item in data_title1:
                    row = i
                    s_cell = "B" + str(row)
                    worksheet1[s_cell].value = item
                    if (i == 15 or i == 22 or i == 29 or i == 31 or i == 35 or i == 38):
                        i += 2
                    else:
                        i += 1
                for item in entry_datas1:
                    row = j
                    s_cell = "C" + str(row)
                    worksheet1[s_cell].value = item
                    if (j == 15 or j == 22 or j == 31):
                        j += 3
                    elif (j == 29 or j == 35 or j == 38):
                        j += 2
                    else:
                        j += 1
                worksheet1['D6'] = "Years"
                write_class(t4b, 8, t4a, 2)
                write_class(t6b, 10, t6a, 2)
                write_class(t7b, 11, t7a, 2)
                write_class(t16, 27, "Years", 1)
                write_class(t23a, 35, "To", 2)
                write_class(t25a, 38, "To", 2)
                write_class(t27a, 41, "To", 2)
                workbook1.save(filename=f'{filename}.xlsx')

            def composite_writer():
                os.chdir(path121)
                filename = "All Teachers"

                try:
                    workbook3 = load_workbook(filename=f'{filename}.xlsx')
                except:
                    workbook3 = Workbook()
                if filename in workbook3.sheetnames:
                    worksheet2 = workbook3[filename]
                else:
                    worksheet2 = workbook3.active
                    worksheet2.title = filename
                if worksheet2.max_row == 1:
                    for column in range(1, 36):
                        for row in range(4, 5):
                            worksheet2.cell(row=row, column=column).value = composite_title[column - 1]

                    for column in range(9, 12):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = location[column - 9]

                    for column in range(12, 15):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = location[column - 12]

                    for column in range(5, 8):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = dob_title[column - 5]

                    for column in range(19, 24):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = education_title[column - 19]
                    for column in range(24, 29):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = experience_title[column - 24]
                    for column in range(30, 36):
                        for row in range(5, 6):
                            worksheet2.cell(row=row, column=column).value = coarses_title[column - 30]

                    composite_data.insert(0, 1)
                    for column in range(1, 36):
                        for row in range(6, 7):
                            worksheet2.cell(row=row, column=column).value = composite_data[column - 1]
                    composite_data.pop(0)
                else:
                    composite_data.insert(0, worksheet2.max_row - 4)
                    row1 = worksheet2.max_row + 1
                    row2 = worksheet2.max_row + 2
                    for column in range(1, 36):
                        for row in range(row1, row2):
                            worksheet2.cell(row=row, column=column).value = composite_data[column - 1]
                    composite_data.pop(0)
                worksheet2['B1'] = "Last Modified On:"
                worksheet2['C1'] = date_time

                workbook3.save(f'{filename}.xlsx')

                text = teacher_name.get() + str(
                    teacher_age.get()) + ".xlsx  File has been created in" + date_time + " \n"
                act_text.config(state='normal')
                act_text.insert("1.0", f'\n {text}')
                act_text.config(state="disabled")

            individual_writer()
            composite_writer()

        def clear_all(event):
            teacher_name_entry.delete(0, "end")
            teacher_age_entry.delete(0, "end")
            teacher_gender_entry.delete(0, "end")
            teacher_dob11.delete(0, "end")
            teacher_dob12.delete(0, "end")
            teacher_dob13.delete(0, "end")
            teacher_mstatus_entry.delete(0, "end")
            teacher_paddress_entry1.delete(0, "end")
            teacher_paddress_entry2.delete(0, "end")
            teacher_paddress_entry3.delete(0, "end")
            teacher_taddress_entry1.delete(0, "end")
            teacher_taddress_entry2.delete(0, "end")
            teacher_taddress_entry3.delete(0, "end")
            teacher_contact_entry.delete(0, "end")
            teacher_s_contact_entry.delete(0, "end")
            teacher_email_entry.delete(0, "end")
            teacher_facebook_entry.delete(0, "end")
            teacher_degree_entry.delete(0, "end")
            teacher_board_entry.delete(0, "end")
            teacher_institution_entry.delete(0, "end")
            teacher_grade_entry.delete(0, "end")
            teacher_subject_entry.delete(0, "end")
            teacher_work_entry.delete(0, "end")
            teacher_post_entry.delete(0, "end")
            teacher_duration_entry.delete(0, "end")
            teacher_winstitution_entry.delete(0, "end")
            teacher_wlocation_entry.delete(0, "end")
            teacher_description_entry.delete('1.0', "end-1c")
            teacher_coarse1_entry.delete(0, "end")
            teacher_c_coarse11_entry.delete(0, "end")
            teacher_c_coarse121_entry.delete(0, "end")
            teacher_coarse2_entry.delete(0, "end")
            teacher_c_coarse211_entry.delete(0, "end")
            teacher_c_coarse221_entry.delete(0, "end")
            teacher_coarse3_entry.delete(0, "end")
            teacher_c_coarse311_entry.delete(0, "end")
            teacher_c_coarse321_entry.delete(0, "end")

        def preview(event):
            teacher_preview1.configure(state='normal')

            teacher_preview1.delete('1.0', "end-1c")
            teacher_preview1.insert(INSERT, f' \t Summary of the Provided Information \n')
            teacher_preview1.insert(INSERT, f' Personal Information \n')
            teacher_preview1.insert(INSERT, f' Name \t  {teacher_name.get()} \n')
            teacher_preview1.insert(INSERT, f' Age \t  {teacher_age.get()} \n')
            teacher_preview1.insert(INSERT, f' Gender \t  {teacher_gender.get()} \n')
            teacher_preview1.insert(INSERT,
                                    f' Date of Birth \t  {teacher_dob1.get()} \t {teacher_dob2.get()} \t {teacher_dob3.get()} \n')
            teacher_preview1.insert(INSERT, f' Marriage Status \t  {teacher_marriage.get()} \n')
            teacher_preview1.insert(INSERT,
                                    f' Permanent Address \t  {teacher_paddress.get()}  \t {teacher_paddress1.get()} \t {teacher_paddress2.get()} \n')
            teacher_preview1.insert(INSERT,
                                    f' Temporary Address \t  {teacher_taddress.get()} \t {teacher_taddress1.get()} \t {teacher_taddress2.get()}\n')
            teacher_preview1.insert(INSERT, f' Contact No. \t  {teacher_contact1.get()} \n')
            teacher_preview1.insert(INSERT, f' Secondary Contact No. \t  {teacher_contact2.get()} \n')
            teacher_preview1.insert(INSERT, f' E-Mail Address \t  {teacher_email.get()} \n')
            teacher_preview1.insert(INSERT, f' Facebook ID \t  {teacher_facebook.get()} \n')
            teacher_preview1.insert(INSERT, f'\n Education \t  \n')
            teacher_preview1.insert(INSERT, f' Level \t  {teacher_level1.get()} \n')
            teacher_preview1.insert(INSERT, f' Board \t  {teacher_board1.get()} \n')
            teacher_preview1.insert(INSERT, f' Institution \t  {teacher_institution1.get()} \n')
            teacher_preview1.insert(INSERT, f' Grade/Percentage \t  {teacher_grade1.get()} \n')
            teacher_preview1.insert(INSERT, f' Main Subject \t  {teacher_subject.get()} \n')
            teacher_preview1.insert(INSERT, f'\n Experiences and Eligibility \t \n')
            teacher_preview1.insert(INSERT, f' Field of Work \t  {teacher_fow1.get()}  \n')
            teacher_preview1.insert(INSERT, f' Post \t  {teacher_post1.get()} \n')
            teacher_preview1.insert(INSERT, f' Duration \t  {teacher_duration1.get()} \t Year/s\n')
            teacher_preview1.insert(INSERT, f' Institution \t  {teacher_wintitution1.get()} \n')
            teacher_preview1.insert(INSERT, f' Location \t  {teacher_wlocation1.get()} \n')
            teacher_preview1.insert(INSERT, f'\t  \n')
            teacher_preview1.insert(INSERT, f' Description \t  {teacher_description} \n')
            teacher_preview1.insert(INSERT, f'\n Teaching Subject Entry \t  \n')
            teacher_preview1.insert(INSERT, f' Main Subject \t  {teacher_mainsubject.get()}')
            teacher_preview1.insert(INSERT, f' For Class/es \t {teacher_class1.get()} ')
            teacher_preview1.insert(INSERT, f' To \t {teacher_class2.get()} \n')
            teacher_preview1.insert(INSERT, f' Secondary Subject1 \t  {teacher_ssubject1.get()} ')
            teacher_preview1.insert(INSERT, f' For Class/es\t {teacher_class21.get()} ')
            teacher_preview1.insert(INSERT, f' To \t {teacher_class22.get()} \n')
            teacher_preview1.insert(INSERT, f' Secondary Subject2 \t  {teacher_ssubject2.get()} ')
            teacher_preview1.insert(INSERT, f' For Class/es \t {teacher_class31.get()} ')
            teacher_preview1.insert(INSERT, f' To \t {teacher_class32.get()} \n')

            teacher_preview1.configure(state='disabled')

        teacher_name = StringVar()
        teacher_age = IntVar()
        teacher_gender = StringVar()
        teacher_dob1 = StringVar()
        teacher_dob2 = StringVar()
        teacher_dob3 = StringVar()
        teacher_marriage = StringVar()
        teacher_paddress = StringVar()
        teacher_paddress1 = StringVar()  #
        teacher_paddress2 = StringVar()  #
        teacher_taddress = StringVar()
        teacher_taddress1 = StringVar()  #
        teacher_taddress2 = StringVar()  #
        teacher_contact1 = StringVar()
        teacher_contact2 = StringVar()
        teacher_email = StringVar()
        teacher_facebook = StringVar()
        # teacher_photo = StringVar()
        teacher_level1 = StringVar()
        teacher_board1 = StringVar()
        teacher_institution1 = StringVar()
        teacher_grade1 = StringVar()
        teacher_subject = StringVar()
        teacher_fow1 = StringVar()
        teacher_post1 = StringVar()
        teacher_duration1 = DoubleVar()
        teacher_wintitution1 = StringVar()
        teacher_wlocation1 = StringVar()
        teacher_description_entry = StringVar()
        teacher_description = StringVar()
        teacher_mainsubject = StringVar()
        teacher_class1 = StringVar()
        teacher_class2 = StringVar()
        teacher_ssubject1 = StringVar()
        teacher_class21 = StringVar()
        teacher_class22 = StringVar()
        teacher_ssubject2 = StringVar()
        teacher_class31 = StringVar()
        teacher_class32 = StringVar()

        options_age = list(range(15, 80))
        options_gender = ("Male", "Female", "Other")
        options_marriage = ("Not Married", "Married")
        options_level = (
            "", "Ph.D", "Masters", "Bachelors", "Diploma", "+2/Higher Education Board", "SLC/SEE", "Class-8")
        options_district = (
            "Bhaktapur", "Chitwan", "Dhading", "Dolakha", "Kathmandu", "Kavrepalanchok", "Lalitpur", "Makwanpur",
            "Nuwakot",
            "Ramechhap", "Rasuwa", "Sindhuli", "Sindhupalchok", "Bhojpur", "Dhankuta", "Ilam", "Jhapa", "Khotang",
            "Morang",
            "Okhaldhunga", "Panchthar", "Sankhuwasabha", "Solukhumbu", "Sunsari", "Taplejung", "Terhathum",
            "Udayapur",
            "Bara",
            "Dhanusa", "Mahottari", "Parsa", "Rautahat", "Saptari", "Sarlahi", "Siraha", "Baglung", "Gorkha",
            "Kaski",
            "Lamjung", "Manang", "Mustang", "Myagdi", "Nawalpur", "Parbat", "Syangja", "Tanahun", "Arghakhanchi",
            "Banke",
            "Bardiya", "Dang", "Eastern Rukum", "Gulmi", "Kapilvastu", "Palpa", "Parasi", "Pyuthan", "Rolpa",
            "Rupandehi",
            "Achham", "Baitadi", "Bajhang", "Bajura", "Dadeldhura", "Darchula", "Doti", "Kailali", "Kanchanpur",
            "Dailekh",
            "Dolpa", "Humla", "Jajarkot", "Jumla", "Kalikot", "Mugu", "Salyan", "Surkhet", "Western Rukum")
        options_board = (
            "Pokhara University", "Tribhuvan University", "Kathmandu Univeristy", "Mid-Western University",
            "Purbanchal University", "Higher Secondary Education Board", "Nepal Education Board")
        options_local_level = (
            "Pokhara Metropolitan City", "Annapurna", "Machhapuchhre", "Madi", "Rupa", "Badigad", "Kathekhola",
            "Nisikhola",
            "Bareng", "Tarakhola", "Tamankhola", "Shahid Lakhan", "Barpak Sulikot", "Aarughat", "Siranchowk",
            "Gandaki",
            "Bhimsen Thapa", "Ajirkot", "Dharche", "Tsum Nubri", "Marsyangdi", "Dordi", "Dudhpokhari",
            "Kwaholasothar",
            "Manang Disyang", "Nason", "Chame", "Narpa Bhumi", "Gharapjhong", "Thasang", "Baragung Muktichhetra",
            "Lomanthang",
            "Lo-Ghekar Damodarkunda", "Malika", "Mangala", "Raghuganga", "Dhaulagiri", "Annapurna", "Hupsekot",
            "Binayi Triveni", "Bulingtar", "Baudikali", "Kaligandaki", "Biruwa", "Harinas", "Aandhikhola",
            "Arjun Chaupari",
            "Phedikhola", "Rishing", "Myagde", "Aanbu Khaireni", "Bandipur", "Ghiring", "Devghat")
        options_subjects = ("Accountancy", "Applied Science", "Applied Finance", "Arts", "Business Administration",
                            "Business, Enterpreneurship and Technology", "Business", "Chemistry", "Commerce",
                            "Economics",
                            "Civil Engineering", "Financial Economics", "Health Administration", "Health Science",
                            "Information Technology", "Law", "Management", "Mathematics", "Medical Science",
                            "Mechanical Engineering", "Medicine", "Pharmaceutical Science", "Physics",
                            "Political Science",
                            "Public Health", "Computer Science", "Nursing", "Social Science", "Urban Planning",
                            "Veterinary Science")
        options_duration = (
            "0", "0.25", "0.5", "0.75", "1", "1.25", "1.5", "1.75", "2", "2.25", "2.5", "2.75", "3", "3.25", "3.5",
            "3.75", "4",
            "4.25", "4.5", "4.75", "5", "5.25", "5.5", "5.75", "6", "6.25", "6.5", "6.75", "7", "7.25", "7.5",
            "7.75",
            "8",
            "8.25", "8.5", "8.75", "9", "9.25", "9.5", "9.75", "10", "10.25", "10.5", "10.75", "11", "11.25",
            "11.5",
            "11.75",
            "12", "12.25", "12.5", "12.75", "13", "13.25", "13.5", "13.75", "14", "14.25", "14.5", "14.75", "15",
            "15.25",
            "15.5", "15.75", "16", "16.25", "16.5", "16.75", "17", "17.25", "17.5", "17.75", "18", "18.25", "18.5",
            "18.75",
            "19", "19.25", "19.5", "19.75", "20", "20.25", "20.5", "20.75", "21", "21.25", "21.5", "21.75", "22",
            "22.25",
            "22.5", "22.75", "23", "23.25", "23.5", "23.75", "24", "24.25", "24.5", "24.75", "25", "25.25", "25.5",
            "25.75",
            "26", "26.25", "26.5", "26.75", "27", "27.25", "27.5", "27.75", "28", "28.25", "28.5", "28.75", "29",
            "29.25",
            "29.5", "29.75", "30", "30.25", "30.5", "30.75")
        options_classes = (
            "Pre-Nursery", "Nursery", "LKG", "UKG", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10")
        options_courses = (
            "Nepali", "English", "Mathematics", "Science", "Social Studies", "General Knowledge", "Health",
            "Population and Environment Science", "Computer Science", "Optional Mathematics", "Occupation",
            "Optional Environmental Science")
        options_year = (
            "2078", "2077", "2076", "2075", "2074", "2073", "2072", "2071", "2070", "2069", "2068", "2067", "2066",
            "2065",
            "2064", "2063", "2062", "2061", "2060", "2059", "2058", "2057", "2056", "2055", "2054", "2053", "2052",
            "2051",
            "2050", "2049", "2048", "2047", "2046", "2045", "2044", "2043", "2042", "2041", "2040", "2039", "2038",
            "2037",
            "2036", "2035", "2034", "2033", "2032", "2031", "2030", "2029", "2028", "2027", "2026", "2025", "2024",
            "2023",
            "2022", "2021", "2020")
        options_month = (
            "Baisakh", "Jestha", "Ashadh", "Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Paush", "Magh",
            "Falgun",
            "Chaitra")
        options_day = (
            "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
            "20", "21",
            "22", "23", "24", "25", "26", "27", "28", "29", "30", "31", "32")

        registration_navigation1 = Frame(detail_pane_frame, relief=FLAT, bg="lightcyan", width=715, height=470)
        registration_navigation1.grid(row=0, column=0, sticky="nsew")

        teacher_frame1 = Frame(registration_navigation1, pady=10, bg="lightcyan")
        teacher_frame1.pack()
        teacher_title_label = Label(teacher_frame1, text="Register - Details of Teacher",
                                    font=("verdana", 12, 'bold'), relief=FLAT, fg="black", bg="coral", width=63)
        teacher_title_label.grid(row=0, columnspan=3, pady=0)

        teacher_frame2 = Frame(registration_navigation1, bg="lightcyan", pady=10)
        teacher_frame2.pack(fill=X)
        teacher_main_label = Label(teacher_frame2, text="Personal Information", font=("verdana", 12, 'bold'),
                                   relief=FLAT, fg="black", anchor=W, bg="lightcyan")
        teacher_main_label.grid(row=0, columnspan=3, padx=5, pady=5, sticky="ew")

        teacher_name_label = Label(teacher_frame2, text="Name", font=("verdana", 10), width=35,
                                   relief=FLAT, fg="black", bg="bisque")
        teacher_name_label.grid(row=1, column=0, pady=5, padx=5)
        teacher_name_entry = Entry(teacher_frame2, font=("verdana", 10), relief=FLAT, fg="black", width=35,
                                   textvariable=teacher_name)
        teacher_name_entry.grid(row=1, column=1, pady=5, padx=5)

        teacher_age_label = Label(teacher_frame2, text="Age", font=("verdana", 10), width=35,
                                  relief=FLAT, fg="black", bg="bisque")
        teacher_age_label.grid(row=2, column=0, pady=5, padx=5)
        teacher_age_entry = ttk.Combobox(teacher_frame2, state="readonly", values=options_age, font=("verdana", 10),
                                         width=33,
                                         textvariable=teacher_age)
        teacher_age_entry.grid(row=2, column=1, pady=5, padx=5)

        teacher_gender_label = Label(teacher_frame2, text="Gender", font=("verdana", 10), width=35,
                                     relief=FLAT, fg="black", bg="bisque")
        teacher_gender_label.grid(row=3, column=0, pady=5, padx=5)
        teacher_gender_entry = ttk.Combobox(teacher_frame2, state="readonly", values=options_gender,
                                            font=("verdana", 10), width=33, textvariable=teacher_gender)
        teacher_gender_entry.grid(row=3, column=1, pady=5, padx=5)

        teacher_dob_label = Label(teacher_frame2, text="Date of Birth (YYYY-MM-DD)", font=("verdana", 10), width=35,
                                  relief=FLAT, fg="black", bg="bisque")
        teacher_dob_label.grid(row=4, column=0, pady=5, padx=5)

        teacher_dob = Frame(teacher_frame2, width=35)
        teacher_dob.grid(row=4, column=1, pady=5)
        teacher_dob11 = ttk.Combobox(teacher_dob, state="readonly", values=options_year,
                                     font=("verdana", 10), width=8, textvariable=teacher_dob1)
        teacher_dob11.pack(side=LEFT)
        teacher_dob12 = ttk.Combobox(teacher_dob, state="readonly", values=options_month,
                                     font=("verdana", 10), width=12, textvariable=teacher_dob2)
        teacher_dob12.pack(side=LEFT)
        teacher_dob13 = ttk.Combobox(teacher_dob, state="readonly", values=options_day,
                                     font=("verdana", 10), width=7, textvariable=teacher_dob3)
        teacher_dob13.pack(side=LEFT)

        teacher_mstatus_label = Label(teacher_frame2, text="Marital Status", font=("verdana", 10), width=35,
                                      relief=FLAT, fg="black", bg="bisque")
        teacher_mstatus_label.grid(row=5, column=0, pady=5, padx=5)
        teacher_mstatus_entry = ttk.Combobox(teacher_frame2, state="readonly", values=options_marriage,
                                             font=("verdana", 10), width=33, textvariable=teacher_marriage)
        teacher_mstatus_entry.grid(row=5, column=1, pady=5, padx=5)

        teacher_paddress_label = Label(teacher_frame2, text="Permanent Address", font=("verdana", 10), width=35,
                                       relief=FLAT, fg="black", bg="bisque")
        teacher_paddress_label.grid(row=6, column=0, pady=5, padx=5)
        teacher_paddress_entry = Frame(teacher_frame2, relief=FLAT, width=35)
        teacher_paddress_entry.grid(row=6, column=1, pady=5, padx=5)
        teacher_paddress_entry1 = ttk.Combobox(teacher_paddress_entry, state="readonly", values=options_district,
                                               font=("verdana", 10),
                                               width=8, textvariable=teacher_paddress)
        teacher_paddress_entry1.current(37)
        teacher_paddress_entry1.pack(side=LEFT)

        teacher_paddress_entry2 = ttk.Combobox(teacher_paddress_entry, state="normal", values=options_local_level,
                                               font=("verdana", 10), width=8, textvariable=teacher_paddress1)
        teacher_paddress_entry2.current(0)
        teacher_paddress_entry2.pack(side=LEFT)
        teacher_paddress_entry3 = Entry(teacher_paddress_entry, font=("verdana", 10), relief=FLAT, fg="black",
                                        width=14, textvariable=teacher_paddress2)
        teacher_paddress_entry3.pack(side=LEFT)

        teacher_taddress_label = Label(teacher_frame2, text="Temporary Address", font=("verdana", 10), width=35,
                                       relief=FLAT, fg="black", bg="bisque")
        teacher_taddress_label.grid(row=7, column=0, pady=5, padx=5)
        teacher_taddress_entry = Frame(teacher_frame2, relief=FLAT, width=35)
        teacher_taddress_entry.grid(row=7, column=1, pady=5, padx=5)
        teacher_taddress_entry1 = ttk.Combobox(teacher_taddress_entry, state="readonly", values=options_district,
                                               font=("verdana", 10),
                                               width=8, textvariable=teacher_taddress)
        teacher_taddress_entry1.current(37)
        teacher_taddress_entry1.pack(side=LEFT)

        teacher_taddress_entry2 = ttk.Combobox(teacher_taddress_entry, state="normal", values=options_local_level,
                                               font=("verdana", 10), width=8, textvariable=teacher_taddress1)
        teacher_taddress_entry2.current(0)
        teacher_taddress_entry2.pack(side=LEFT)
        teacher_taddress_entry3 = Entry(teacher_taddress_entry, font=("verdana", 10), relief=FLAT, fg="black",
                                        width=14, textvariable=teacher_taddress2)
        teacher_taddress_entry3.pack(side=LEFT)

        teacher_contact_label = Label(teacher_frame2, text="Contact Number", font=("verdana", 10), width=35,
                                      relief=FLAT, fg="black", bg="bisque")
        teacher_contact_label.grid(row=8, column=0, pady=5, padx=5)
        teacher_contact_entry = Entry(teacher_frame2, font=("verdana", 10), relief=FLAT, fg="black", width=35,
                                      textvariable=teacher_contact1)
        teacher_contact_entry.grid(row=8, column=1, pady=5, padx=5)

        teacher_s_contact_label = Label(teacher_frame2, text=" Secondary Contact Number *", font=("verdana", 10),
                                        width=35,
                                        relief=FLAT, fg="black", bg="bisque")
        teacher_s_contact_label.grid(row=9, column=0, pady=5, padx=5)
        teacher_s_contact_entry = Entry(teacher_frame2, font=("verdana", 10), relief=FLAT, fg="black", width=35,
                                        textvariable=teacher_contact2)
        teacher_s_contact_entry.grid(row=9, column=1, pady=5, padx=5)

        teacher_email_label = Label(teacher_frame2, text="E-mail Address", font=("verdana", 10), width=35,
                                    relief=FLAT, fg="black", bg="bisque")
        teacher_email_label.grid(row=10, column=0, pady=5, padx=5)
        teacher_email_entry = Entry(teacher_frame2, font=("verdana", 10), relief=FLAT, fg="black", width=35,
                                    textvariable=teacher_email)
        teacher_email_entry.grid(row=10, column=1, pady=5, padx=5)

        teacher_facebook_label = Label(teacher_frame2, text="Facebook ID *", font=("verdana", 10), width=35,
                                       relief=FLAT, fg="black", bg="bisque")
        teacher_facebook_label.grid(row=11, column=0, pady=5, padx=5)
        teacher_facebook_entry = Entry(teacher_frame2, font=("verdana", 10), relief=FLAT, fg="black", width=35,
                                       textvariable=teacher_facebook)
        teacher_facebook_entry.grid(row=11, column=1, pady=5, padx=5)

        teacher_photo_label = Label(teacher_frame2, text="Photo (Drop ___.png Extension here)",
                                    font=("verdana", 10),
                                    width=35,
                                    relief=FLAT, fg="black", bg="bisque")
        teacher_photo_label.grid(row=12, column=0, pady=5, padx=5)
        teacher_photo_entry = Entry(teacher_frame2, font=("verdana", 10, 'bold'), relief=FLAT, fg="black", width=35)
        teacher_photo_entry.grid(row=12, column=1, pady=5, padx=5)

        teacher_frame3 = Frame(registration_navigation1, bg="lightcyan", pady=10)
        teacher_frame3.pack(fill=X)
        teacher_education_label = Label(teacher_frame3, text="Academic Qualification", font=("verdana", 12, 'bold'),
                                        relief=FLAT, fg="black", anchor=W, bg="lightcyan")
        teacher_education_label.grid(row=0, columnspan=4, padx=5, pady=5, sticky="ew")

        teacher_degree_label = Label(teacher_frame3, text="Level/Degree", font=("verdana", 10), width=15,
                                     relief=FLAT, fg="black", bg="bisque")
        teacher_degree_label.grid(row=1, column=0, pady=5, padx=5)
        teacher_degree_entry = ttk.Combobox(teacher_frame3, state="readonly", values=options_level,
                                            font=("verdana", 10), width=15,
                                            textvariable=teacher_level1)
        teacher_degree_entry.current(2)
        teacher_degree_entry.grid(row=2, column=0, pady=5, padx=5)

        teacher_board_label = Label(teacher_frame3, text="University/Board", font=("verdana", 10), width=19,
                                    relief=FLAT, fg="black", bg="bisque")
        teacher_board_label.grid(row=1, column=1, pady=5, padx=5)
        teacher_board_entry = ttk.Combobox(teacher_frame3, state="normal", values=options_board,
                                           font=("verdana", 10), width=17, textvariable=teacher_board1)
        teacher_board_entry.current(2)
        teacher_board_entry.grid(row=2, column=1, pady=5, padx=5)

        teacher_institution_label = Label(teacher_frame3, text="Institution", font=("verdana", 10), width=19,
                                          relief=FLAT, fg="black", bg="bisque")
        teacher_institution_label.grid(row=1, column=2, pady=5, padx=5)
        teacher_institution_entry = Entry(teacher_frame3, font=("verdana", 10), relief=FLAT, fg="black", width=17,
                                          textvariable=teacher_institution1)
        teacher_institution_entry.grid(row=2, column=2, pady=5, padx=5)
        teacher_grade_label = Label(teacher_frame3, text="Percentage/Grades", font=("verdana", 10), width=19,
                                    relief=FLAT, fg="black", bg="bisque")
        teacher_grade_label.grid(row=1, column=3, pady=5, padx=5)
        teacher_grade_entry = Entry(teacher_frame3, font=("verdana", 10), relief=FLAT, fg="black", width=17,
                                    textvariable=teacher_grade1)
        teacher_grade_entry.grid(row=2, column=3, pady=5, padx=5)

        teacher_subject_label = Label(teacher_frame3, text="Major Subject", font=("verdana", 10), width=19,
                                      relief=FLAT, fg="black", bg="bisque")
        teacher_subject_label.grid(row=3, column=0, pady=5, padx=5)
        teacher_subject_entry = ttk.Combobox(teacher_frame3, state="normal", values=options_subjects,
                                             font=("verdana", 10), width=8, textvariable=teacher_subject)
        teacher_subject_entry.grid(row=3, column=1, pady=5, padx=5)

        teacher_frame4 = Frame(registration_navigation1, bg="lightcyan", pady=10)
        teacher_frame4.pack(fill=X)
        teacher_experience_label = Label(teacher_frame4, text="Experiences and Eligibility",
                                         font=("verdana", 12, 'bold'),
                                         relief=FLAT, fg="black", anchor=W, bg="lightcyan")
        teacher_experience_label.grid(row=0, columnspan=5, padx=5, pady=5, sticky="ew")

        teacher_work_label = Label(teacher_frame4, text="Field of Work", font=("verdana", 10), width=15,
                                   relief=FLAT, fg="black", bg="bisque")
        teacher_work_label.grid(row=1, column=0, pady=5, padx=5)
        teacher_work_entry = Entry(teacher_frame4, font=("verdana", 10), relief=FLAT, fg="black", width=14,
                                   textvariable=teacher_fow1)
        teacher_work_entry.grid(row=2, column=0, pady=5, padx=5)

        teacher_post_label = Label(teacher_frame4, text="Post", font=("verdana", 10), width=15,
                                   relief=FLAT, fg="black", bg="bisque")
        teacher_post_label.grid(row=1, column=1, pady=5, padx=5)
        teacher_post_entry = Entry(teacher_frame4, font=("verdana", 10), relief=FLAT, fg="black", width=14,
                                   textvariable=teacher_post1)
        teacher_post_entry.grid(row=2, column=1, pady=5, padx=5)

        teacher_duration_label = Label(teacher_frame4, text="Duration", font=("verdana", 10), width=15,
                                       relief=FLAT, fg="black", bg="bisque")
        teacher_duration_label.grid(row=1, column=2, pady=5, padx=5)
        teacher_duration_entry = ttk.Combobox(teacher_frame4, state="readonly", values=options_duration,
                                              font=("verdana", 10), width=8, textvariable=teacher_duration1)
        teacher_duration_entry.grid(row=2, column=2, pady=5, padx=5)

        teacher_winstitution_label = Label(teacher_frame4, text="Institution", font=("verdana", 10), width=15,
                                           relief=FLAT, fg="black", bg="bisque")
        teacher_winstitution_label.grid(row=1, column=3, pady=5, padx=5)
        teacher_winstitution_entry = Entry(teacher_frame4, font=("verdana", 10), relief=FLAT, fg="black", width=14,
                                           textvariable=teacher_wintitution1)
        teacher_winstitution_entry.grid(row=2, column=3, pady=5, padx=5)

        teacher_wlocation_label = Label(teacher_frame4, text="Location of Work", font=("verdana", 10), width=15,
                                        relief=FLAT, fg="black", bg="bisque")
        teacher_wlocation_label.grid(row=1, column=4, pady=5, padx=5)
        teacher_wlocation_entry = Entry(teacher_frame4, font=("verdana", 10), relief=FLAT, fg="black", width=14,
                                        textvariable=teacher_wlocation1)
        teacher_wlocation_entry.grid(row=2, column=4, pady=5, padx=5)

        teacher_frame5 = Frame(registration_navigation1, bg="lightcyan", pady=10)
        teacher_frame5.pack(fill=X)
        teacher_experience_label = Label(teacher_frame5, text="Description of the Teacher",
                                         font=("verdana", 12, 'bold'),
                                         fg="black", anchor=W, bg="lightcyan")
        teacher_experience_label.grid(row=0, columnspan=5, sticky="w")

        teacher_frame6 = Frame(registration_navigation1, bg="lightcyan", width=40, height=100)
        teacher_frame6.pack(fill=X)
        teacher_description_entry = scrolledtext.ScrolledText(teacher_frame6, font=("verdana", 10), relief=FLAT,
                                                              fg="black", width=60,
                                                              height=5)
        teacher_description_entry.grid(row=0, column=0, columnspan=4, rowspan=4)
        teacher_description = teacher_description_entry.get('1.0', "end-1c")

        teacher_frame7 = Frame(registration_navigation1, bg="lightcyan", pady=10)
        teacher_frame7.pack(fill=X)
        teacher_education_label = Label(teacher_frame7, text="Coarse Registration", font=("verdana", 12, 'bold'),
                                        relief=FLAT, fg="black", anchor=W, bg="lightcyan")
        teacher_education_label.grid(row=0, columnspan=4, padx=5, pady=5, sticky="ew")

        teacher_coarse1_label = Label(teacher_frame7, text="Main Subject", font=("verdana", 10), width=20,
                                      relief=FLAT, fg="black", bg="bisque")
        teacher_coarse1_label.grid(row=1, column=0, pady=5, padx=5)
        teacher_coarse1_entry = ttk.Combobox(teacher_frame7, state="readonly", values=options_courses,
                                             font=("verdana", 10), width=15, textvariable=teacher_mainsubject)
        teacher_coarse1_entry.grid(row=1, column=1, pady=5, padx=5)
        teacher_c_coarse1_label = Label(teacher_frame7, text="Class Range (A-Z)", font=("verdana", 10), width=19,
                                        relief=FLAT, fg="black", bg="bisque")
        teacher_c_coarse1_label.grid(row=1, column=2, pady=5, padx=5)

        teacher_c_coarse1_entry = Frame(teacher_frame7, relief=FLAT)
        teacher_c_coarse1_entry.grid(row=1, column=3, pady=5, padx=5)
        teacher_c_coarse11_entry = ttk.Combobox(teacher_c_coarse1_entry, state="readonly", values=options_classes,
                                                font=("verdana", 10), width=6, textvariable=teacher_class1)
        teacher_c_coarse11_entry.pack(side=LEFT)
        teacher_c_coarse12_label = Label(teacher_c_coarse1_entry, text="-", font=("verdana", 10, 'bold'), width=1,
                                         relief=FLAT, fg="black", bg="lightcyan")
        teacher_c_coarse12_label.pack(side=LEFT)

        teacher_c_coarse121_entry = ttk.Combobox(teacher_c_coarse1_entry, state="readonly", values=options_classes,
                                                 font=("verdana", 10), width=6, textvariable=teacher_class2)
        teacher_c_coarse121_entry.pack(side=LEFT)

        teacher_coarse2_label = Label(teacher_frame7, text="Secondary Subject-1 *", font=("verdana", 10), width=20,
                                      relief=FLAT, fg="black", bg="bisque")
        teacher_coarse2_label.grid(row=2, column=0, pady=5, padx=5)
        teacher_coarse2_entry = ttk.Combobox(teacher_frame7, state="readonly", values=options_courses,
                                             font=("verdana", 10), width=15, textvariable=teacher_ssubject1)
        teacher_coarse2_entry.grid(row=2, column=1, pady=5, padx=5)
        teacher_c_coarse2_label = Label(teacher_frame7, text="Class Range (A-Z)", font=("verdana", 10), width=19,
                                        relief=FLAT, fg="black", bg="bisque")
        teacher_c_coarse2_label.grid(row=2, column=2, pady=5, padx=5)

        teacher_c_coarse2_entry = Frame(teacher_frame7, relief=FLAT)
        teacher_c_coarse2_entry.grid(row=2, column=3, pady=5, padx=5)
        teacher_c_coarse211_entry = ttk.Combobox(teacher_c_coarse2_entry, state="readonly", values=options_classes,
                                                 font=("verdana", 10), width=6, textvariable=teacher_class21)
        teacher_c_coarse211_entry.pack(side=LEFT)
        teacher_c_coarse212_label = Label(teacher_c_coarse2_entry, text="-", font=("verdana", 10, 'bold'), width=1,
                                          relief=FLAT, fg="black", bg="lightcyan")
        teacher_c_coarse212_label.pack(side=LEFT)
        teacher_c_coarse221_entry = ttk.Combobox(teacher_c_coarse2_entry, state="readonly", values=options_classes,
                                                 font=("verdana", 10), width=6, textvariable=teacher_class22)
        teacher_c_coarse221_entry.pack(side=LEFT)

        teacher_coarse3_label = Label(teacher_frame7, text="Secondary Subject-2 *", font=("verdana", 10), width=20,
                                      relief=FLAT, fg="black", bg="bisque")
        teacher_coarse3_label.grid(row=3, column=0, pady=5, padx=5)
        teacher_coarse3_entry = ttk.Combobox(teacher_frame7, state="normal", values=options_courses,
                                             font=("verdana", 10), width=15, textvariable=teacher_ssubject2)
        teacher_coarse3_entry.grid(row=3, column=1, pady=5, padx=5)
        teacher_c_coarse3_label = Label(teacher_frame7, text="Class Range (A-Z)", font=("verdana", 10), width=19,
                                        relief=FLAT, fg="black", bg="bisque")
        teacher_c_coarse3_label.grid(row=3, column=2, pady=5, padx=5)

        teacher_c_coarse3_entry = Frame(teacher_frame7, relief=FLAT)
        teacher_c_coarse3_entry.grid(row=3, column=3, pady=5, padx=5)

        teacher_c_coarse311_entry = ttk.Combobox(teacher_c_coarse3_entry, state="normal", values=options_classes,
                                                 font=("verdana", 10), width=6, textvariable=teacher_class31)
        teacher_c_coarse311_entry.pack(side=LEFT)
        teacher_c_coarse312_label = Label(teacher_c_coarse3_entry, text="-", font=("verdana", 10, 'bold'), width=1,
                                          relief=FLAT, fg="black", bg="lightcyan")
        teacher_c_coarse312_label.pack(side=LEFT)

        teacher_c_coarse321_entry = ttk.Combobox(teacher_c_coarse3_entry, state="readonly", values=options_classes,
                                                 font=("verdana", 10), width=6, textvariable=teacher_class32)
        teacher_c_coarse321_entry.pack(side=LEFT)

        teacher_preview1 = Frame(registration_navigation1, bg="lightcyan", pady=10, width=60, height=100)
        teacher_preview1.pack(fill=X)
        teacher_preview1.rowconfigure(0, weight=1)
        teacher_preview1.columnconfigure(0, weight=1)

        teacher_preview1 = scrolledtext.ScrolledText(teacher_preview1, font=("verdana", 10), relief=FLAT,
                                                     fg="black",
                                                     width=80, height=10)
        teacher_preview1.grid(row=0, column=0, columnspan=4, rowspan=4)

        teacher_frame8 = Frame(registration_navigation1, bg="lightcyan", pady=10)
        teacher_frame8.pack(fill=X)
        teacher_note_label = Label(teacher_frame8, text="-Data are optional to the fields with asterick(*) sign",
                                   font=("verdana", 10, 'bold'),
                                   relief=FLAT, fg="black", anchor=W, bg="lightcyan")
        teacher_note_label.grid(row=0, columnspan=3, padx=5, pady=5, sticky="ew")

        teacher_button_preview = Button(teacher_frame8, text="Preview Summary", font=("verdana", 12), relief=RAISED,
                                        bg="springgreen2", width=20)
        teacher_button_preview.bind("<Button-1>", preview)
        teacher_button_preview.grid(row=1, column=0, pady=5, padx=10)

        teacher_button_submit = Button(teacher_frame8, text="Submit", font=("verdana", 12), relief=RAISED,
                                       bg="springgreen2", width=20)
        teacher_button_submit.bind("<Button-1>", teacher_to_db)
        teacher_button_submit.grid(row=1, column=1, pady=5, padx=10)

        teacher_button_clear = Button(teacher_frame8, text="Clear All", font=("verdana", 12), relief=RAISED,
                                      bg="lightpink", width=20)
        teacher_button_clear.bind("<Button-1>", clear_all)
        teacher_button_clear.grid(row=1, column=2, pady=5, padx=10)

    def existing_record_event1(event):
        existing_navigation1 = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
        existing_navigation1.grid(row=0, column=0, sticky="nsew")

        details = Button(existing_navigation1, text="Details of the Teacher", font=("verdana", 12), width=24,
                         fg="black", bg="bisque", pady=5, border=1, relief=GROOVE)
        details.grid(row=0, column=0, sticky="ew")

    def teachers_event(event):
        teachers_navigation = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
        teachers_navigation.grid(row=0, column=0, sticky="nsew")

        new_registration1 = Button(teachers_navigation, text="Registration", font=("verdana", 12),
                                   fg="black", bg="bisque", pady=5, border=1, relief=GROOVE, width=24)
        new_registration1.bind("<Button-1>", new_registration_event1)
        new_registration1.grid(row=0, column=0, sticky="we")

        existing_record1 = Button(teachers_navigation, text="Existing Record", font=("verdana", 12),
                                  fg="black", bg="bisque", pady=5, border=1, relief=GROOVE, width=24)
        existing_record1.bind("<Button-1>", existing_record_event1)
        existing_record1.grid(row=1, column=0, sticky="we")

    def new_registration_event2(event):
        pass

    def existing_record_event2(event):
        existing_navigation2 = Frame(navigation_pane_frame, width=300, height=470, relief=FLAT, bg="white")
        existing_navigation2.place(x=0, y=0)

        details1 = Button(existing_navigation2, text="Details of the Staff", font=("verdana", 12), width=24,
                          relief=FLAT,
                          fg="black", bg="bisque")
        details1.pack(side=TOP, anchor=NW, expand=1, fill=X, padx=0, pady=5)

    def staffs_event(event):
        staffs_navigation = Frame(navigation_pane_frame, width=250, height=470, relief=FLAT)
        staffs_navigation.place(x=0, y=0)
        canvas3 = Canvas(staffs_navigation, width=250, height=470)
        canvas3.pack()

        new_registration2 = Button(canvas3, text="Registration", font=("verdana", 12), width=24, relief=FLAT,
                                   fg="black", bg="bisque")
        new_registration2.bind("<Button-1>", new_registration_event2)
        new_registration2.pack(side=TOP, anchor=NW, expand=1, fill=X, padx=0, pady=5)

        existing_record2 = Button(canvas3, text="Existing Record", font=("verdana", 12), width=24, relief=FLAT,
                                  fg="black", bg="bisque")
        existing_record2.bind("<Button-1>", existing_record_event2)
        existing_record2.pack(side=TOP, anchor=NW, expand=1, fill=X, padx=0, pady=5)


    myframe = Frame(root, bg='turquoise', relief=GROOVE)
    myframe.place(x=10, y=10)

    # ------------ Scroll Bar Configuration
    canvas = Canvas(myframe)
    frame = Frame(canvas)
    myscrollbar = Scrollbar(myframe, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=myscrollbar.set)

    myscrollbar.pack(side="right", fill="y")
    canvas.pack(side="left")
    canvas.create_window((0, 0), window=frame, anchor='nw')
    frame.bind("<Configure>", myfunction_button1)

    # ------------ Sub Windows Assignment
    top_frame = Frame(frame, bg='turquoise', height=600, bd=5)
    top_frame.pack(side=TOP, fill=X)
    menu_frame = Frame(frame)
    menu_frame.pack(side=TOP)
    pane_frame = Frame(frame, bd=5, relief=FLAT, width=1300, height=470)
    pane_frame.pack(side=TOP)
    bottom_frame = Frame(frame, bd=0)
    bottom_frame.pack(side=BOTTOM, anchor=W)

    navigation_pane_frame = Frame(pane_frame, width=250, height=470, relief=FLAT)
    navigation_pane_frame.place(x=0, y=0)
    navigation_pane_frame.rowconfigure(0, weight=1)
    navigation_pane_frame.columnconfigure(0, weight=1)

    space_pane_frame = Frame(pane_frame, width=5, height=470)
    space_pane_frame.place(x=250, y=0)
    action_pane_frame = Frame(pane_frame, bg="purple", width=270, height=470)
    action_pane_frame.place(x=970, y=0)
    action_pane_frame.rowconfigure(0, weight=1)
    action_pane_frame.columnconfigure(0, weight=1)

    action_widgets = Frame(action_pane_frame, bg="orange", width=270, height=40)
    action_widgets.grid(row=0, column=0, sticky='nsew')
    action_progress = Frame(action_pane_frame, bg="purple", width=270, height=430)
    action_progress.grid(row=1, column=0, sticky='nsew')
    act_scroll = Scrollbar(action_progress)
    act_text = Text(action_progress, wrap='word', font=('verdana', 10), state='disabled', width=32,
                    yscrollcommand=act_scroll.set)
    act_scroll.configure(command=act_text.yview)
    act_text.grid(row=0, column=0, sticky='nwes')
    act_scroll.grid(row=0, column=1, sticky="ns")

    detail_pane_frame1 = Frame(pane_frame, bg="lightcyan", width=715, height=460)
    detail_pane_frame1.place(x=255, y=0)
    maincanvas = Canvas(detail_pane_frame1, width=695, height=460)
    detail_pane_frame2 = Frame(maincanvas, bg="green")
    myscrollbar1 = Scrollbar(detail_pane_frame1, orient="vertical", command=maincanvas.yview)
    maincanvas.configure(yscrollcommand=myscrollbar1.set)
    myscrollbar1.pack(side=RIGHT, fill=Y)
    maincanvas.pack(side=LEFT)
    maincanvas.create_window((0, 0), window=detail_pane_frame2, anchor='nw')
    detail_pane_frame2.bind("<Configure>", myfunction_button11)
    detail_pane_frame = Frame(detail_pane_frame2, relief=FLAT, bg="lightcyan", width=715, height=460)
    detail_pane_frame.pack(fill=BOTH)
    detail_pane_frame.rowconfigure(0, weight=1)
    detail_pane_frame.columnconfigure(0, weight=1)

    # ------------Gadgets and Labels
    welcome_text = Label(top_frame, text=f'Welcome to the School Management System of {school_name}',
                         font=('Tahoma', 20, 'bold'), fg="Crimson")
    welcome_text.pack(side=TOP, fill=X)
    # welcome_image1 = Label(top_frame, image=welcome_image)
    # welcome_image1.pack()
    # ------------
    students_database = Button(menu_frame, text="Students Database", font=("verdana", 12, "bold"), relief=GROOVE)
    students_database.bind("<Button-1>", students_event)
    students_database.pack(side=LEFT, padx=0)

    teachers_database = Button(menu_frame, text="Teachers Database", font=("verdana", 12, "bold"), relief=GROOVE)
    teachers_database.bind("<Button-1>", teachers_event)
    teachers_database.pack(side=LEFT, padx=100)

    staffs_database = Button(menu_frame, text=f'Staffs Database', font=("verdana", 12, "bold"), relief=GROOVE)
    staffs_database.bind("<Button-1>", staffs_event)
    staffs_database.pack(side=LEFT, padx=0)

    date_label = Label(bottom_frame, text=f"{datetime.datetime.now().strftime('%Y:%m:%d')}",
                       font=("verdana", 12, "italic"))
    date_label.pack(side=LEFT)

    date_label = Label(bottom_frame, text=" / ", font=("verdana", 14, "bold"))
    date_label.pack(side=LEFT)

    obj1 = Clock(bottom_frame)


def directory():
    if os.path.exists(main_folder) is False:
        os.makedirs(main_folder)
        os.chdir(path1)
        os.makedirs(folder1)
        os.makedirs(folder2)
        os.makedirs(folder3)
        os.makedirs(folder4)
        os.makedirs(folder_reg)

        os.chdir(path11)
        os.makedirs(folder11)
        os.makedirs(folder12)
        os.makedirs(folder13)
        os.makedirs(folder14)

        os.chdir(path112)
        os.makedirs(folder12pnur)
        os.makedirs(folder12nur)
        os.makedirs(folder12lkg)
        os.makedirs(folder12ukg)
        os.makedirs(folder121)
        os.makedirs(folder122)
        os.makedirs(folder123)
        os.makedirs(folder124)
        os.makedirs(folder125)
        os.makedirs(folder126)
        os.makedirs(folder127)
        os.makedirs(folder128)
        os.makedirs(folder129)
        os.makedirs(folder1210)

        os.chdir(path12)
        os.makedirs(folder21)
        os.makedirs(folder22)

        os.chdir(path13)
        os.makedirs(folder31)
        os.makedirs(folder32)


directory()
framework(root)
root.mainloop()

############ Programming Code Finished########
